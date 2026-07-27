"""日报导出真实 golden：以 07-07 定稿为模板续列，逐单元格对齐 07-08 定稿。"""
from __future__ import annotations

import os
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.pipeline.export.daily_exporter import export_daily
from scripts.run_chain_regression import load_expected


def _mentor_root() -> Path:
    value = os.getenv("HR_REAL_DAILY_ROOT")
    if not value:
        pytest.skip("未显式配置 mentor-local 真实日报目录")
    return Path(value).expanduser().resolve()


def test_appended_daily_matches_finalized_cell_by_cell(tmp_path):
    root = _mentor_root()
    template = root / "2026-07-07" / "员工数增减情况日报-7月_20260707.xlsx"
    final = root / "2026-07-08" / "员工数增减情况日报-7月_20260708.xlsx"
    if not (template.exists() and final.exists()):
        pytest.skip("真实定稿日报不存在")
    report_date = date(2026, 7, 8)
    expected_rows, tenure_rows = load_expected(final, report_date)
    ctx = {
        "report_date": report_date,
        "rows": {n: {"value": v} for n, v in expected_rows.items()},
    }
    tenure = {
        "rows": tenure_rows,
        "total": {
            "business_unit": "合计",
            "ytd_leavers": sum(r["ytd_leavers"] for r in tenure_rows),
            "avg_tenure_years": None,
        },
    }

    generated_path = export_daily(
        ctx, tenure, str(tmp_path), template_path=str(template),
    )

    generated = load_workbook(generated_path)
    expected = load_workbook(final)
    assert generated.sheetnames == expected.sheetnames == ["Sheet1", "在岗时长"]
    for sheet_name in expected.sheetnames:
        generated_sheet = generated[sheet_name]
        expected_sheet = expected[sheet_name]
        assert (generated_sheet.max_row, generated_sheet.max_column) == (
            expected_sheet.max_row,
            expected_sheet.max_column,
        ), sheet_name
        assert list(generated_sheet.values) == list(expected_sheet.values), sheet_name
    # 新列样式继承：表头日期格式与前一列一致
    sheet1 = generated["Sheet1"]
    assert sheet1.cell(1, sheet1.max_column).number_format == \
        sheet1.cell(1, sheet1.max_column - 1).number_format


_PERSIST_ROWS = (
    2, 3, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 30, 33,
    38, 39, 40,
)


def test_hr_confirmed_month_opening_matches_finalized_july_first(tmp_path):
    """月初 exporter golden：HR 已确认的 B 列覆盖 → 对齐 07-01 定稿。

    这里的 baseline_values 明确代表 month_opening_baselines 中 HR 上传/确认的
    独立月初事实，不再声称可以从 06-30 daily_reports 自动推导重述值。"""
    root = _mentor_root()
    june_final = root / "2026-06-30" / "员工数增减情况日报-6月_20260630.xlsx"
    july_first = root / "2026-07-01" / "员工数增减情况日报-7月_20260701.xlsx"
    if not (june_final.exists() and july_first.exists()):
        pytest.skip("真实定稿日报不存在")
    report_date = date(2026, 7, 1)
    expected_rows, tenure_rows = load_expected(july_first, report_date)
    finalized = load_workbook(july_first)
    baseline_values = {
        n: finalized["Sheet1"].cell(n, 2).value
        for n in _PERSIST_ROWS
        if finalized["Sheet1"].cell(n, 2).value is not None
    }
    ctx = {
        "report_date": report_date,
        "rows": {n: {"value": v} for n, v in expected_rows.items()},
    }
    tenure = {
        "rows": tenure_rows,
        "total": {
            "business_unit": "合计",
            "ytd_leavers": sum(r["ytd_leavers"] for r in tenure_rows),
            "avg_tenure_years": None,
        },
    }

    generated_path = export_daily(
        ctx, tenure, str(tmp_path),
        baseline_values=baseline_values,
        template_path=str(june_final),
    )

    generated = load_workbook(generated_path)
    assert generated.sheetnames == finalized.sheetnames == ["Sheet1", "在岗时长"]
    style_diffs = []
    for sheet_name in finalized.sheetnames:
        generated_sheet = generated[sheet_name]
        expected_sheet = finalized[sheet_name]
        assert (generated_sheet.max_row, generated_sheet.max_column) == (
            expected_sheet.max_row,
            expected_sheet.max_column,
        ), sheet_name
        assert list(generated_sheet.values) == list(expected_sheet.values), sheet_name
        for row in expected_sheet.iter_rows():
            for cell in row:
                mine = generated_sheet.cell(cell.row, cell.column)
                if (mine.number_format != cell.number_format
                        or bool(mine.font.b) != bool(cell.font.b)
                        or mine.fill.fgColor.rgb != cell.fill.fgColor.rgb):
                    style_diffs.append((sheet_name, cell.row, cell.column))
    assert style_diffs == []
