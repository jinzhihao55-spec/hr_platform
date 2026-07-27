"""月初基线必须由 HR 显式确认或上传，不得从上月末静默猜测重述值。"""
from __future__ import annotations

import asyncio
from datetime import date, datetime
from pathlib import Path

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook


def _make_opening_workbook(path: Path, baseline_date: date) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "事项")
    ws.cell(1, 2, datetime.combine(baseline_date, datetime.min.time()))
    ws.cell(1, 2).number_format = "yyyy-mm-dd"
    for row in range(2, 41):
        ws.cell(row, 1, f"Row{row}")
        ws.cell(row, 2, 0)
    values = {
        8: 73, 9: 59, 13: 159, 14: 123, 15: 0, 16: 0, 17: 36,
        30: 14, 37: 73, 38: 0, 39: 0, 40: 73, 18: 73,
    }
    for row, value in values.items():
        ws.cell(row, 2, value)

    tenure = wb.create_sheet("在岗时长")
    tenure.append([
        "事业部", "YTD离职人数", "平均在职（年）",
        datetime.combine(baseline_date, datetime.min.time()),
    ])
    counts = [20, 18, 17, 16, 15, 14, 13, 10]
    for idx, (letter, count) in enumerate(zip("ABCDEFGH", counts), start=2):
        tenure.append([f"BU_{letter}", count, 1.5, None])
        tenure.cell(idx, 2).number_format = "#,##0"
    tenure.append(["合计", sum(counts), None, None])
    wb.save(path)
    return path


def test_upload_persists_hr_month_opening_baseline(db, tmp_path, monkeypatch):
    """上传的 B 列与在岗时长是独立月初事实，不能写回 06-30 日报。"""
    from app.config import settings
    from app.services import month_opening_service

    monkeypatch.setattr(settings, "output_dir", str(tmp_path / "outputs"))
    source = _make_opening_workbook(tmp_path / "opening.xlsx", date(2026, 6, 30))
    with source.open("rb") as stream:
        result = asyncio.run(month_opening_service.import_baseline(
            db,
            report_month=date(2026, 7, 1),
            confirmed_by="hr-user",
            file=UploadFile(filename=source.name, file=stream),
        ))

    assert result["status"] == "confirmed"
    assert result["source_type"] == "uploaded"
    stored = month_opening_service.get_confirmed(db, date(2026, 7, 15))
    assert stored["baseline_rows"][13] == 159
    assert stored["baseline_rows"][14] == 123
    assert stored["baseline_rows"][17] == 36
    assert stored["baseline_rows"][40] == 73
    assert sum(row["ytd_leavers"] for row in stored["tenure_rows"]) == 123
    template = load_workbook(stored["template_path"])
    assert template["Sheet1"].max_column == 2

    generation = month_opening_service.prepare_generation(db, date(2026, 7, 1))
    assert generation["baseline_date"] == date(2026, 6, 30)
    assert generation["baseline_override"][13] == 159
    assert sum(
        row["ytd_leavers"] for row in generation["tenure_baseline"]["rows"]
    ) == 123


def test_month_opening_is_required_before_cross_month_generation(db, tmp_path, monkeypatch):
    from app.config import settings
    from app.core.exceptions import MonthOpeningBaselineMissingError
    from app.services import month_opening_service

    monkeypatch.setattr(settings, "output_dir", str(tmp_path / "outputs"))
    try:
        month_opening_service.require_for_report(db, date(2026, 7, 1))
    except MonthOpeningBaselineMissingError as exc:
        assert exc.code == "month_opening_baseline_missing"
    else:
        raise AssertionError("缺月初基线时必须阻断")


def test_carry_forward_requires_explicit_baseline_date(db, tmp_path, monkeypatch):
    """HR 可明确确认沿用指定定稿；系统不能自行挑一个文件当作已确认。"""
    from app.config import settings
    from app.services import month_opening_service

    output_dir = tmp_path / "outputs"
    finalized = output_dir / "finalized"
    finalized.mkdir(parents=True)
    source = _make_opening_workbook(
        finalized / "员工数增减情况日报_2026-06-30.xlsx",
        date(2026, 6, 30),
    )
    monkeypatch.setattr(settings, "output_dir", str(output_dir))

    result = month_opening_service.confirm_carry_forward(
        db,
        report_month=date(2026, 7, 1),
        baseline_date=date(2026, 6, 30),
        confirmed_by="hr-user",
    )

    assert result["source_type"] == "carry_forward"
    assert result["baseline_date"] == "2026-06-30"
    assert Path(result["template_path"]).is_file()


def test_failed_update_preserves_previous_month_opening_artifact(db, tmp_path, monkeypatch):
    """数据库更新失败不能删掉上一版仍有效的月初模板。"""
    from app.config import settings
    from app.services import month_opening_service

    output_dir = tmp_path / "outputs"
    monkeypatch.setattr(settings, "output_dir", str(output_dir))
    destination = output_dir / "month_opening" / "month_opening_2026-07.xlsx"
    destination.parent.mkdir(parents=True)
    destination.write_bytes(b"previous-approved-artifact")
    prepared = _make_opening_workbook(tmp_path / "replacement.xlsx", date(2026, 6, 30))
    monkeypatch.setattr(
        month_opening_service.month_opening_repo,
        "save",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("db down")),
    )

    try:
        month_opening_service._persist(
            db,
            report_month=date(2026, 7, 1),
            baseline_date=date(2026, 6, 30),
            source_type="uploaded",
            baseline_rows={8: 1, 9: 1, 13: 1, 14: 1, 30: 1},
            tenure_rows=[],
            confirmed_by="hr-user",
            prepared_artifact=prepared,
        )
    except RuntimeError:
        pass
    else:
        raise AssertionError("数据库失败必须向调用方返回错误")

    assert destination.read_bytes() == b"previous-approved-artifact"
