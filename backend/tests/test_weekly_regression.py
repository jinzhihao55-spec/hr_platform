"""Weekly regression checks against the locally archived HR source files."""
from __future__ import annotations

import os
from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.pipeline.calculation.weekly import compute_weekly
from app.pipeline.calculation import validators
from app.pipeline.cleansing import cleanse
from app.pipeline.export.weekly_exporter import export_weekly
from app.pipeline.input import parsers
from app.pipeline.input.daily_workbook import parse_daily_workbook
from app.models.reports import DailyReport
from app.repositories import input_repo, report_repo


def _mentor_root() -> Path:
    value = os.getenv("HR_REAL_DAILY_ROOT")
    if not value:
        pytest.skip("未显式配置 mentor-local 真实日报目录")
    return Path(value).expanduser().resolve()


def _source(day: date, label: str, suffix: str) -> Path:
    folder = _mentor_root() / day.isoformat()
    return folder / f"{label}_{day.strftime('%Y%m%d')}.{suffix}"


def _ingest_week(db, start: date, end: date) -> None:
    day = start
    while day <= end:
        employees_path = _source(day, "人员表", "xls")
        resignations_path = _source(day, "离职人员报表", "xls")
        if not employees_path.exists() or not resignations_path.exists():
            pytest.skip(f"真实周报回归源文件不存在：{day}")

        employees = parsers.parse_employees(str(employees_path))
        employees = cleanse.normalize_employee_types(employees)
        employees, _ = cleanse.filter_inclusion(employees)
        employees = cleanse.trace_resign_transfer(employees)
        input_repo.upsert_employees(db, employees, day)

        resignations = parsers.parse_resignations(str(resignations_path))
        input_repo.upsert_resignations(db, resignations)
        db.commit()
        day += timedelta(days=1)


def test_weekly_export_matches_finalized_workbook_values_and_layout(db, tmp_path):
    final_weekly = (
        _mentor_root()
        / "2026-07-10"
        / "员工数增减周报-7月_W28_20260710.xlsx"
    )
    if not final_weekly.exists():
        pytest.skip("真实定稿周报不存在")
    start = date(2026, 7, 6)
    end = date(2026, 7, 10)
    _ingest_week(db, start, end)

    generated_path = export_weekly(compute_weekly(db, start, end), str(tmp_path))
    generated = load_workbook(generated_path, data_only=True)
    expected = load_workbook(final_weekly, data_only=True)

    assert generated.sheetnames == expected.sheetnames == ["Sheet2", "Sheet1"]
    for sheet_name in expected.sheetnames:
        actual_sheet = generated[sheet_name]
        expected_sheet = expected[sheet_name]
        assert (actual_sheet.max_row, actual_sheet.max_column) == (
            expected_sheet.max_row,
            expected_sheet.max_column,
        )
        assert list(actual_sheet.values) == list(expected_sheet.values)
        assert {str(rng) for rng in actual_sheet.merged_cells.ranges} == {
            str(rng) for rng in expected_sheet.merged_cells.ranges
        }


def test_weekly_validation_blocks_daily_reconciliation_mismatch(db):
    start = date(2026, 7, 6)
    end = date(2026, 7, 10)
    _ingest_week(db, start, end)
    for offset, onboard, resign in [
        (0, 4, 0), (1, 4, 0), (2, 4, 1), (3, 4, 0), (4, 5, 2),
    ]:
        db.add(DailyReport(
            report_date=start + timedelta(days=offset),
            daily_onboard=onboard,
            daily_resign=resign,
        ))
    db.commit()

    checks = validators.run_weekly_checks(compute_weekly(db, start, end))
    hard = validators.hard_failures(checks)

    assert any(check["check"] == "周报本周入职=日报Row2合计" for check in hard)


def test_weekly_validation_blocks_missing_window_end_snapshot():
    checks = validators.run_weekly_checks({
        "week_end": date(2026, 7, 10),
        "snapshot_found": False,
        "main_rows": [],
        "cc_rows": [],
        "daily_reconciliation": {"available_days": 0},
    })

    assert any(
        check["check"] == "周报窗口末人员快照存在"
        for check in validators.hard_failures(checks)
    )


def test_weekly_validation_does_not_compare_incomplete_daily_window():
    checks = validators.run_weekly_checks({
        "week_end": date(2026, 7, 10),
        "snapshot_found": True,
        "main_rows": [{
            "business_unit": "NENT",
            "headcount": 1,
            "cnt_formal": 1,
            "cnt_intern": 0,
            "cnt_labor": 0,
            "joiners": 3,
            "leavers": 0,
        }],
        "cc_rows": [],
        "daily_reconciliation": {
            "available_days": 1,
            "complete": False,
            "joiners": 1,
            "leavers": 0,
        },
    })

    assert not any(
        check["check"] == "周报本周入职=日报Row2合计"
        for check in validators.hard_failures(checks)
    )


def test_weekly_validation_blocks_unconfigured_sheet1_rows():
    checks = validators.run_weekly_checks({
        "week_end": date(2026, 7, 10),
        "snapshot_found": True,
        "main_rows": [],
        "cc_rows": [{
            "cost_center": "9999",
            "project": "临时项目",
            "headcount": 1,
            "joiners": 0,
            "leavers": 0,
        }],
        "daily_reconciliation": {"complete": False},
    })

    assert any(
        check["check"] == "Sheet1项目族与配置一致"
        for check in validators.hard_failures(checks)
    )


def test_weekly_validations_pass_finalized_daily_window(db):
    root = _mentor_root()
    start = date(2026, 7, 6)
    end = date(2026, 7, 10)
    _ingest_week(db, start, end)
    day = start
    while day <= end:
        workbook = (
            root
            / day.isoformat()
            / f"员工数增减情况日报-7月_{day.strftime('%Y%m%d')}.xlsx"
        )
        if not workbook.exists():
            pytest.skip(f"真实定稿日报不存在：{day}")
        rows, _ = parse_daily_workbook(workbook, day)
        report_repo.save_daily(db, day, rows)
        day += timedelta(days=1)

    checks = validators.run_weekly_checks(compute_weekly(db, start, end))

    assert validators.hard_failures(checks) == []
