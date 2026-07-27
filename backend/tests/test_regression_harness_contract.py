"""Release-regression harness keeps all HR artifacts outside Git."""

import json
import os
import shutil
import subprocess
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from scripts.run_single_user_regression import (
    RELEASE_CHECKS,
    REPO_ROOT,
    compare_workbooks_redacted,
    main as regression_main,
    run_fake_regression,
    run_regression,
)
from scripts.scan_sensitive_artifacts import main as privacy_main, scan_paths


def test_harness_requires_external_input_and_output_directories(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="outside repository"):
        run_regression(REPO_ROOT / "tests", tmp_path)


def test_privacy_scanner_reports_rule_without_echoing_secret(tmp_path: Path) -> None:
    secret = "FAKE-SENSITIVE-VALUE"
    trace = tmp_path / "trace.json"
    trace.write_text(f'{{"证件号":"{secret}"}}', encoding="utf-8")

    findings = scan_paths(
        [tmp_path], sensitive_values={"certificate_number": {secret}}
    )

    assert findings[0].path == trace
    assert findings[0].rule == "certificate_number"
    assert secret not in findings[0].message


def test_fake_mode_exercises_every_release_artifact_check(tmp_path: Path) -> None:
    summary = run_fake_regression(tmp_path)

    assert summary["status"] == "passed"
    assert summary["checks"] == {check: True for check in RELEASE_CHECKS}


def test_workbook_comparison_never_returns_cell_values(tmp_path: Path) -> None:
    expected_path = tmp_path / "expected.xlsx"
    actual_path = tmp_path / "actual.xlsx"
    expected = Workbook()
    expected.active["B2"] = "SENSITIVE-EXPECTED"
    expected.active["B2"].font = Font(bold=True)
    expected.active.merge_cells("A3:B3")
    expected.save(expected_path)
    actual = Workbook()
    actual.active["B2"] = "SENSITIVE-ACTUAL"
    actual.save(actual_path)

    result = compare_workbooks_redacted(actual_path, expected_path)

    assert result["value_mismatches"] == ["Sheet!B2"]
    assert "Sheet!B2" in result["style_mismatches"]
    assert result["merge_mismatches"] == ["Sheet"]
    serialized = str(result)
    assert "SENSITIVE-EXPECTED" not in serialized
    assert "SENSITIVE-ACTUAL" not in serialized


def test_workbook_comparison_uses_style_semantics_not_internal_ids(
    tmp_path: Path,
) -> None:
    expected_path = tmp_path / "semantic-expected.xlsx"
    actual_path = tmp_path / "semantic-actual.xlsx"
    expected = Workbook()
    expected.active["B2"] = "same"
    expected.active["B2"].font = Font(bold=True)
    expected.save(expected_path)
    actual = Workbook()
    actual.active["A1"].font = Font(italic=True)
    actual.active["B2"] = "same"
    actual.active["B2"].font = Font(bold=True)
    actual.save(actual_path)

    result = compare_workbooks_redacted(actual_path, expected_path)

    assert "Sheet!A1" in result["style_mismatches"]
    assert "Sheet!B2" not in result["style_mismatches"]


def test_mentor_failure_writes_only_a_redacted_summary(tmp_path: Path) -> None:
    input_root = tmp_path / "mentor-input"
    output_root = tmp_path / "mentor-output"
    for report_date in (date(2026, 7, 7), date(2026, 7, 8)):
        folder = input_root / report_date.isoformat()
        folder.mkdir(parents=True)
        compact = report_date.strftime("%Y%m%d")
        for stem, suffix in (
            ("人员表", ".xls"),
            ("离职人员报表", ".xls"),
            ("协议签署", ".png"),
            ("招聘数据", ".png"),
            ("员工数增减情况日报", ".xlsx"),
        ):
            (folder / f"{stem}_SENSITIVE-FILENAME_{compact}{suffix}").write_bytes(
                b"malformed"
            )
        if report_date == date(2026, 7, 8):
            (
                folder
                / f"员工数增减周报_SENSITIVE-FILENAME_{compact}.xlsx"
            ).write_bytes(b"malformed")

    summary = run_regression(
        input_root,
        output_root,
        start=date(2026, 7, 8),
    )

    assert summary["status"] == "blocked"
    assert summary["weekly_cases"] == [
        {
            "week_end": "2026-07-08",
            "status": "blocked",
            "error_code": "weekly_input_or_calculation_failed",
        }
    ]
    assert [path.name for path in output_root.iterdir()] == [
        "regression_summary.json"
    ]
    serialized = (output_root / "regression_summary.json").read_text(
        encoding="utf-8"
    )
    json.loads(serialized)
    assert "SENSITIVE-FILENAME" not in serialized
    assert str(input_root) not in serialized


def test_mentor_replay_preserves_earlier_evidence_when_a_later_day_blocks(
    tmp_path: Path, monkeypatch
) -> None:
    input_root = tmp_path / "accepted-input"
    output_root = tmp_path / "accepted-output"
    for report_date in (
        date(2026, 7, 7),
        date(2026, 7, 8),
        date(2026, 7, 9),
    ):
        folder = input_root / report_date.isoformat()
        folder.mkdir(parents=True)
        workbook = Workbook()
        workbook.active.title = "Sheet1"
        workbook.active["A1"] = "事项"
        workbook.active["B1"] = report_date
        workbook.save(
            folder / f"员工数增减情况日报_测试_{report_date:%Y%m%d}.xlsx"
        )

    calls = []

    def replay(args):
        calls.append([value.isoformat() for value in args.dates])
        current = args.dates[-1]
        if current == date(2026, 7, 9):
            raise RuntimeError("SENSITIVE-PIPELINE-DETAIL")
        expected = next(
            (input_root / current.isoformat()).glob("员工数增减情况日报*.xlsx")
        )
        args.output_dir.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(
            expected,
            args.output_dir / f"员工数增减情况日报_{current.isoformat()}.xlsx",
        )
        return {
            "baseline_tenure_match": True,
            "cases": [
                {
                    "report_date": current.isoformat(),
                    "passed": True,
                    "row_mismatches": [],
                    "tenure_total_match": True,
                    "tenure_row_mismatches": [],
                    "hard_failures": [],
                }
            ],
        }

    monkeypatch.setattr(
        "scripts.run_single_user_regression.run_chain_regression.run", replay
    )

    summary = run_regression(
        input_root,
        output_root,
        start=date(2026, 7, 8),
    )

    assert calls == [["2026-07-08"], ["2026-07-08", "2026-07-09"]]
    assert [case["status"] for case in summary["cases"]] == [
        "passed",
        "blocked",
    ]
    assert "SENSITIVE-PIPELINE-DETAIL" not in json.dumps(summary)


def test_mentor_calculation_failure_is_not_masked_as_missing_output(
    tmp_path: Path, monkeypatch
) -> None:
    input_root = tmp_path / "accepted-input"
    output_root = tmp_path / "accepted-output"
    for report_date in (date(2026, 7, 7), date(2026, 7, 8)):
        folder = input_root / report_date.isoformat()
        folder.mkdir(parents=True)
        workbook = Workbook()
        workbook.active.title = "Sheet1"
        workbook.active["A1"] = "事项"
        workbook.active["B1"] = report_date
        workbook.save(
            folder / f"员工数增减情况日报_测试_{report_date:%Y%m%d}.xlsx"
        )

    def replay(args):
        return {
            "baseline_tenure_match": True,
            "cases": [
                {
                    "report_date": args.dates[-1].isoformat(),
                    "passed": False,
                    "row_mismatches": [{"row": 2, "kind": "value_mismatch"}],
                    "tenure_total_match": True,
                    "tenure_row_mismatches": [],
                    "hard_failures": [],
                }
            ],
        }

    monkeypatch.setattr(
        "scripts.run_single_user_regression.run_chain_regression.run", replay
    )

    summary = run_regression(
        input_root,
        output_root,
        start=date(2026, 7, 8),
    )

    assert summary["status"] == "failed"
    assert summary["error_code"] == "daily_regression_failed"
    assert summary["cases"] == [
        {
            "report_date": "2026-07-08",
            "status": "failed",
            "checks": {"calculation": False},
            "row_mismatches": [{"row": 2, "kind": "value_mismatch"}],
            "tenure_row_mismatches": [],
            "hard_failures": [],
            "workbook_mismatches": {},
        }
    ]


def test_fake_cli_runs_without_an_output_directory(capsys) -> None:
    assert regression_main(["--mode", "fake"]) == 0
    assert capsys.readouterr().out.strip() == "fake PASS checks=12/12"


def test_fake_cli_suppresses_value_bearing_pipeline_logs() -> None:
    environment = {**os.environ, "PYTHONPATH": str(REPO_ROOT)}
    result = subprocess.run(
        [
            sys.executable,
            str(REPO_ROOT / "scripts/run_single_user_regression.py"),
            "--mode",
            "fake",
        ],
        cwd=REPO_ROOT,
        env=environment,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert result.stdout.strip() == "fake PASS checks=12/12"
    assert result.stderr == ""


def test_privacy_cli_detects_key_material_without_printing_it(
    tmp_path: Path, capsys
) -> None:
    marker = "-" * 5 + "BEGIN PRIVATE KEY" + "-" * 5
    (tmp_path / "unsafe.log").write_text(marker, encoding="utf-8")

    assert privacy_main([str(tmp_path)]) == 1
    output = capsys.readouterr().out
    assert "private_key" in output
    assert marker not in output
