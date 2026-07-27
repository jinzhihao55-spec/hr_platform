"""日报导出版式契约（合成数据，CI 可跑）：对齐人工定稿格式。

定稿事实（project/日报 实测）：
- 工作簿只有 Sheet1 + 在岗时长；Sheet1 按日追加日期列，Row25/29/36 为分节日期戳；
- 月初文件从「报告日前最近一份已验收定稿」派生：A 列标签/样式承袭，
  上月末列成为 B 基线列（值以库内已验收基线覆盖），再追加当月首日列；
- 找不到任何可承袭的工作簿时必须阻断，禁止凭空造模板；
- 在岗时长 sheet 原地刷新（值更新、样式保留），报告日仅在 D1。"""
from datetime import date, datetime

import pytest
from openpyxl import Workbook, load_workbook

from app.core.exceptions import DailyTemplateMissingError
from app.pipeline.export.daily_exporter import export_daily


def _ctx(report_date: date, onboard: int = 3) -> dict:
    return {
        "report_date": report_date,
        "rows": {2: {"value": onboard, "label": "今日入职"},
                 3: {"value": 1, "label": "今日离职"}},
    }


def _tenure(leavers: int = 0) -> dict:
    rows = [{"slot": f"BU_{c}", "business_unit": f"BU_{c}",
             "ytd_leavers": leavers, "avg_tenure_years": None}
            for c in "ABCDEFGH"]
    return {"rows": rows,
            "total": {"business_unit": "合计", "ytd_leavers": leavers * 8,
                      "avg_tenure_years": None}}


def _make_template(path, *dates: date) -> str:
    """手工造一个最小模板：A 标签 + 每个日期一列；在岗时长 sheet 带样式标记。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "事项")
    for idx, day in enumerate(dates, start=2):
        c = ws.cell(1, idx, datetime(day.year, day.month, day.day))
        c.number_format = "yyyy-mm-dd"
    ws.cell(2, 1, "今日入职")
    ws.cell(26, 1, "WICRESOFT VIET NAM")
    for idx in range(2, len(dates) + 2):
        ws.cell(2, idx, 9)
        ws.cell(26, idx, 0)
        for row in (25, 29, 36):
            stamp = ws.cell(row, idx, ws.cell(1, idx).value)
            stamp.number_format = "yyyy-mm-dd"
    ws.cell(40, 1, "尾行")
    tenure_sheet = wb.create_sheet("在岗时长")
    tenure_sheet.append(["事业部", "YTD离职人数", "平均在职（年）",
                         datetime(dates[-1].year, dates[-1].month, dates[-1].day)])
    for c in "ABCDEFGH":
        tenure_sheet.append([f"BU_{c}", 1, None, None])
    tenure_sheet.append(["合计", 8, None, None])
    tenure_sheet.cell(2, 2).number_format = "#,##0"   # 样式标记：原地刷新必须保留
    out = str(path)
    wb.save(out)
    return out


def test_export_without_any_template_is_blocked(tmp_path):
    with pytest.raises(DailyTemplateMissingError):
        export_daily(_ctx(date(2026, 7, 1)), _tenure(), str(tmp_path))


def test_same_month_template_appends_one_column(tmp_path):
    template = _make_template(tmp_path / "t.xlsx", date(2026, 7, 1))
    path = export_daily(_ctx(date(2026, 7, 2), onboard=5), _tenure(2),
                        str(tmp_path), template_path=template)

    workbook = load_workbook(path)
    sheet1 = workbook["Sheet1"]
    assert workbook.sheetnames == ["Sheet1", "在岗时长"]
    assert sheet1.max_column == 3
    assert sheet1.cell(1, 3).value == datetime(2026, 7, 2)
    assert sheet1.cell(2, 3).value == 5
    assert sheet1.cell(2, 2).value == 9          # 历史列原样
    for row in (25, 29, 36):
        assert sheet1.cell(row, 3).value == datetime(2026, 7, 2)
    tenure_sheet = workbook["在岗时长"]
    assert tenure_sheet.cell(1, 4).value == datetime(2026, 7, 2)
    assert tenure_sheet.cell(2, 2).value == 2               # 值已刷新
    assert tenure_sheet.cell(2, 2).number_format == "#,##0"  # 样式保留（原地刷新）


def test_cross_month_template_derives_month_start(tmp_path):
    template = _make_template(
        tmp_path / "june.xlsx", date(2026, 6, 29), date(2026, 6, 30),
    )
    path = export_daily(
        _ctx(date(2026, 7, 1), onboard=4), _tenure(1), str(tmp_path),
        baseline_values={2: 7},           # 库内已验收基线覆盖 B 列
        template_path=template,
    )

    sheet1 = load_workbook(path)["Sheet1"]
    assert sheet1.max_column == 3                    # 事项 + 6/30 基线 + 7/1
    assert sheet1.cell(1, 2).value == datetime(2026, 6, 30)
    assert sheet1.cell(1, 3).value == datetime(2026, 7, 1)
    assert sheet1.cell(26, 1).value == "WICRESOFT VIET NAM"  # 标签承袭定稿
    assert sheet1.cell(2, 2).value == 7              # B 列被库内基线覆盖
    assert sheet1.cell(26, 2).value == 0             # 未覆盖行保留上月末列值
    assert sheet1.cell(2, 3).value == 4
    assert sheet1.cell(25, 2).value == datetime(2026, 6, 30)  # 分节戳保留基线日
