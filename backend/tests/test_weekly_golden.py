"""周报合成 golden 端到端：入库 → 计算 → 硬校验 → 导出 → 重读断言。

真实数据回归（test_weekly_regression.py）依赖本机绝对路径、CI 会跳过；
本用例用合成数据在 CI 覆盖同一条链路的结构与口径契约。"""
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook

from app.models.reports import DailyReport
from app.pipeline.calculation import validators
from app.pipeline.calculation.weekly import compute_weekly
from app.pipeline.export.weekly_exporter import export_weekly
from app.repositories import input_repo


MON, FRI = date(2026, 7, 6), date(2026, 7, 10)
AIA = "友邦保险系统平台开发及优化项目"


def _emp(no, bu, etype="正式员工", proj="内部项目X", status="在职",
         entry=date(2025, 1, 1), resign=None):
    return {"工号": no, "中文名": no, "员工类型": etype, "员工状态": status,
            "入职日期": entry, "离职日期": resign, "事业部": bu, "事业部编号": bu,
            "项目编号": proj, "项目名称": proj}


def _seed_week(db):
    baseline = [
        _emp("A1", "NINS", proj=AIA),
        _emp("A2", "NINS", etype="劳务人员", proj=f"{AIA}-运维"),
        _emp("B1", "NENT"),
        _emp("L1", "NENT", proj="内部项目Y"),
    ]
    input_repo.upsert_employees(db, pd.DataFrame(baseline), MON)
    db.commit()

    friday = [
        _emp("A1", "NINS", proj=AIA),
        _emp("A2", "NINS", etype="劳务人员", proj=f"{AIA}-运维"),
        _emp("B1", "NENT"),
        _emp("L1", "NENT", proj="内部项目Y", status="离职", resign=date(2026, 7, 8)),
        _emp("J1", "NENT", etype="长期实习生", entry=date(2026, 7, 8)),
    ]
    input_repo.upsert_employees(db, pd.DataFrame(friday), FRI)
    db.commit()

    # 完整日报窗口：周三 1 入 1 离，其余为 0 —— 与上面事实一致
    for offset in range(5):
        day = MON + timedelta(days=offset)
        db.add(DailyReport(
            report_date=day,
            daily_onboard=1 if day == date(2026, 7, 8) else 0,
            daily_resign=1 if day == date(2026, 7, 8) else 0,
        ))
    db.commit()


def test_weekly_golden_chain_passes_hard_checks_and_exports(db, tmp_path):
    _seed_week(db)

    ctx = compute_weekly(db, MON, FRI)
    assert validators.hard_failures(validators.run_weekly_checks(ctx)) == []

    rows = {r["business_unit"]: r for r in ctx["main_rows"]}
    assert rows["NINS"]["headcount"] == 2
    assert (rows["NINS"]["cnt_formal"], rows["NINS"]["cnt_labor"]) == (1, 1)
    # 项目族归并：友邦两个变体合并为一个族，计数 2
    assert rows["NINS"]["top3_projects"][0] == {"name": AIA, "count": 2}
    assert rows["NENT"]["headcount"] == 2
    assert (rows["NENT"]["joiners"], rows["NENT"]["leavers"]) == (1, 1)

    path = export_weekly(ctx, str(tmp_path))
    workbook = load_workbook(path, data_only=True)
    assert workbook.sheetnames == ["Sheet2", "Sheet1"]

    sheet2 = workbook["Sheet2"]
    values = list(sheet2.values)
    # 主体列合并且填竖排标签
    assert values[2][0] == "微\n创\n网\n络"
    total_row = values[-1]
    assert total_row[0] == "合计"
    assert total_row[2:8] == (4, 2, 1, 1, 1, 1)

    sheet1 = workbook["Sheet1"]
    s1_values = list(sheet1.values)
    assert len(s1_values) == 1 + 6  # 表头 + 固定 6 个项目族
    assert s1_values[1][:3] == ("9000", AIA, 2)


def test_weekly_export_matches_the_approved_style_contract(db, tmp_path):
    _seed_week(db)
    workbook = load_workbook(
        export_weekly(compute_weekly(db, MON, FRI), str(tmp_path))
    )
    sheet2 = workbook["Sheet2"]

    assert sheet2["A1"].fill.fgColor.rgb == "00DDEBF7"
    assert sheet2["D1"].fill.fgColor.rgb == "00E2F0D9"
    assert sheet2["G1"].fill.fgColor.rgb == "00FCE4D6"
    assert sheet2["I1"].fill.fgColor.rgb == "00FFFF00"
    assert sheet2["A3"].font.name == "Microsoft YaHei"
    assert sheet2["A3"].font.sz == 10
    assert sheet2["A3"].alignment.horizontal == "center"
    assert sheet2["I3"].alignment.horizontal == "left"
    assert {
        sheet2["A3"].border.left.style,
        sheet2["A3"].border.right.style,
        sheet2["A3"].border.top.style,
        sheet2["A3"].border.bottom.style,
    } == {"thin"}
    total = sheet2.cell(sheet2.max_row, 1)
    assert total.font.bold is True
    assert total.font.color.rgb == "00FF0000"
    assert all(
        sheet2.row_dimensions[row].height == 24
        for row in range(1, sheet2.max_row + 1)
    )
    assert [sheet2.column_dimensions[column].width for column in "ABCDEFGHIJ"] == [
        8,
        12,
        12,
        12,
        12,
        12,
        14,
        14,
        38,
        12,
    ]

    sheet1 = workbook["Sheet1"]
    assert sheet1["A2"].alignment.horizontal == "center"
    assert sheet1["B2"].alignment.horizontal == "left"
    assert all(
        sheet1.row_dimensions[row].height == 24
        for row in range(1, sheet1.max_row + 1)
    )
    assert [sheet1.column_dimensions[column].width for column in "ABCDE"] == [
        12,
        34,
        12,
        14,
        14,
    ]
