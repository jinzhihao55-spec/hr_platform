"""Publication writes verified artifacts and compatibility projections atomically."""

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.domain.fact_bundle import FactBundle
from app.models.publication import PublicationAttempt, PublishedReport, ReportArtifact
from app.models.reports import DailyReport, WeeklyReport
from app.models.runs import ReportRun, RunReportTarget, RunStatus, TargetStatus
from app.pipeline.calculation.daily import ITEMS
from app.pipeline.input.daily_workbook import parse_daily_workbook
from app.repositories import publication_repo
from app.services import publication_service
from app.services.preview_service import build_preview
from app.services.publication_service import (
    PreviewChanged,
    PublicationFailed,
    PublicationRecoveryRequired,
    publish,
    recover_publication_attempts,
)
from tests.test_preview_service import (
    make_daily_bundle,
    make_run,
    make_weekly_bundle,
)


def _daily_template(path: Path) -> str:
    baseline_date = date(2026, 7, 7)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "事项")
    sheet.cell(1, 2, datetime(2026, 7, 7))
    for row in range(2, 41):
        sheet.cell(row, 1, ITEMS.get(row, f"保留行{row}"))
        sheet.cell(row, 2, 0)
    for row in (25, 29, 36):
        sheet.cell(row, 2, datetime(2026, 7, 7))

    tenure = workbook.create_sheet("在岗时长")
    tenure.append(["事业部", "YTD离职人数", "平均在职（年）", datetime(2026, 7, 7)])
    from app.core import constants as C

    for label in C.get_tenure_bu_labels():
        tenure.append([label, 0, None, None])
    tenure.append(["合计", 0, None, None])
    workbook.save(path)
    assert baseline_date.isoformat() in str(path) or path.is_file()
    return str(path)


def _previous_daily(db, report_date: date) -> PublishedReport:
    previous_run = ReportRun(
        report_date=report_date,
        status=RunStatus.ready.value,
        rule_version="rules-v0",
    )
    db.add(previous_run)
    db.flush()
    previous = PublishedReport(
        run_id=previous_run.id,
        report_kind="daily",
        period_start=report_date,
        period_end=report_date,
        version=1,
        is_current=True,
        snapshot_json="{}",
        snapshot_hash="a" * 64,
        published_by="local-operator",
        published_at=datetime(2026, 7, 8, 9, 0, 0),
    )
    db.add(previous)
    db.add(DailyReport(report_date=report_date, daily_onboard=7))
    db.commit()
    return previous


def test_publication_error_redaction_drops_exception_payload():
    secret = "SECRET-SNAPSHOT-CONTENT"

    message = publication_service._redacted_error(ValueError(secret))

    assert secret not in message
    assert "ValueError" in message


def test_export_failure_leaves_previous_report_and_projection_current(
    db, tmp_path, monkeypatch
):
    report_date = date(2026, 7, 8)
    previous = _previous_daily(db, report_date)
    run = make_run(db, report_date)
    bundle = make_daily_bundle(report_date)
    build_preview(db, run.id, "daily", bundle=bundle)

    def fail_write(*_args, **_kwargs):
        raise OSError("disk full")

    monkeypatch.setattr(publication_service, "_write_artifacts", fail_write)

    with pytest.raises(PublicationFailed, match="文件写入失败"):
        publish(
            db,
            run.id,
            ["daily"],
            "local-operator",
            bundles={"daily": bundle},
            output_dir=tmp_path,
        )

    assert publication_repo.current_daily(db, report_date).id == previous.id
    assert db.scalar(select(DailyReport).where(DailyReport.report_date == report_date)).daily_onboard == 7
    assert db.scalar(select(func.count()).select_from(PublishedReport)) == 1


def test_daily_publish_reloads_excel_and_updates_projection(db, tmp_path):
    run = make_run(db, date(2026, 7, 8))
    bundle = make_daily_bundle()
    preview = build_preview(db, run.id, "daily", bundle=bundle)
    template = _daily_template(tmp_path / "日报模板_2026-07-07.xlsx")

    reports = publish(
        db,
        run.id,
        ["daily"],
        "local-operator",
        bundles={"daily": bundle},
        template_paths={"daily": template},
        output_dir=tmp_path,
    )

    report = reports[0]
    assert report.snapshot_hash == preview.snapshot_hash
    assert publication_repo.current_daily(db, date(2026, 7, 8)).id == report.id
    projection = db.scalar(
        select(DailyReport).where(DailyReport.report_date == date(2026, 7, 8))
    )
    assert projection.daily_onboard == 1
    artifacts = db.scalars(
        select(ReportArtifact).where(ReportArtifact.report_id == report.id)
    ).all()
    assert {artifact.artifact_kind for artifact in artifacts} == {
        "excel",
        "execution_log",
        "event_ledger",
        "validation_report",
        "manifest",
    }
    assert all(Path(artifact.protected_path).is_file() for artifact in artifacts)
    excel = next(artifact for artifact in artifacts if artifact.artifact_kind == "excel")
    parsed, parsed_date = parse_daily_workbook(
        Path(excel.protected_path), date(2026, 7, 8)
    )
    assert parsed_date == date(2026, 7, 8)
    assert parsed[2]["value"] == preview.rows[2].value
    target = db.scalar(
        select(RunReportTarget).where(
            RunReportTarget.run_id == run.id,
            RunReportTarget.report_kind == "daily",
        )
    )
    assert target.status == TargetStatus.published.value
    assert target.published_report_id == report.id


def test_weekly_publish_reloads_excel_and_updates_projection(db, tmp_path):
    run = make_run(db, date(2026, 7, 10))
    bundle = make_weekly_bundle()
    preview = build_preview(
        db,
        run.id,
        "weekly",
        bundle=bundle,
        week_start=date(2026, 7, 6),
        week_end=date(2026, 7, 10),
    )

    report = publish(
        db,
        run.id,
        ["weekly"],
        "local-operator",
        bundles={"weekly": bundle},
        periods={"weekly": (date(2026, 7, 6), date(2026, 7, 10))},
        output_dir=tmp_path,
    )[0]

    assert report.snapshot_hash == preview.snapshot_hash
    assert publication_repo.current_weekly(
        db, date(2026, 7, 6), date(2026, 7, 10)
    ).id == report.id
    projection = db.scalar(select(WeeklyReport))
    assert projection.headcount_active == 1
    excel = db.scalar(
        select(ReportArtifact).where(
            ReportArtifact.report_id == report.id,
            ReportArtifact.artifact_kind == "excel",
        )
    )
    assert Path(excel.protected_path).is_file()


def test_republishing_same_run_is_rejected_before_staging(db, tmp_path):
    run = make_run(db, date(2026, 7, 10))
    bundle = make_weekly_bundle()
    period = (date(2026, 7, 6), date(2026, 7, 10))
    build_preview(
        db,
        run.id,
        "weekly",
        bundle=bundle,
        week_start=period[0],
        week_end=period[1],
    )
    publish(
        db,
        run.id,
        ["weekly"],
        "local-operator",
        bundles={"weekly": bundle},
        periods={"weekly": period},
        output_dir=tmp_path,
    )
    attempt_count = db.scalar(select(func.count()).select_from(PublicationAttempt))

    with pytest.raises(PublicationFailed, match="create a revision"):
        publish(
            db,
            run.id,
            ["weekly"],
            "local-operator",
            bundles={"weekly": bundle},
            periods={"weekly": period},
            output_dir=tmp_path,
        )

    assert db.scalar(select(func.count()).select_from(PublicationAttempt)) == attempt_count


def test_changed_facts_require_a_new_preview(db, tmp_path):
    run = make_run(db, date(2026, 7, 8))
    previewed = make_daily_bundle()
    build_preview(db, run.id, "daily", bundle=previewed)
    changed = FactBundle(
        report_date=previewed.report_date,
        baseline_date=previewed.baseline_date,
        rule_version=previewed.rule_version,
        employments=previewed.employments.iloc[0:0],
        baseline_rows=previewed.baseline_rows,
        daily_reconciliation=previewed.daily_reconciliation,
    )

    with pytest.raises(PreviewChanged):
        publish(
            db,
            run.id,
            ["daily"],
            "local-operator",
            bundles={"daily": changed},
            output_dir=tmp_path,
        )

    assert db.scalar(select(func.count()).select_from(PublishedReport)) == 0
    target = db.scalar(
        select(RunReportTarget).where(
            RunReportTarget.run_id == run.id,
            RunReportTarget.report_kind == "daily",
        )
    )
    assert target.preview_hash is None
    assert target.status == TargetStatus.needs_review.value


def test_recovery_finishes_metadata_committed_directory_move(db, tmp_path):
    run = make_run(db, date(2026, 7, 8))
    report = PublishedReport(
        run_id=run.id,
        report_kind="daily",
        period_start=run.report_date,
        period_end=run.report_date,
        version=1,
        is_current=True,
        snapshot_json="{}",
        snapshot_hash="b" * 64,
        published_by="local-operator",
        published_at=datetime(2026, 7, 8, 9, 0, 0),
    )
    db.add(report)
    db.flush()
    staging = tmp_path / ".publication-staging" / "attempt"
    final = tmp_path / "published" / "daily" / "2026-07-08" / "attempt"
    staging.mkdir(parents=True)
    (staging / "验证报告.json").write_text("{}", encoding="utf-8")
    attempt = PublicationAttempt(
        run_id=run.id,
        report_kind="daily",
        status="metadata_committed",
        staging_path=str(staging),
        final_path=str(final),
        report_id=report.id,
    )
    db.add(attempt)
    db.commit()

    recovered = recover_publication_attempts(db, tmp_path)

    assert recovered == [attempt.id]
    assert final.joinpath("验证报告.json").is_file()
    assert not staging.exists()
    db.refresh(attempt)
    assert attempt.status == "completed"


def test_ambiguous_metadata_commit_preserves_staging_for_recovery(
    db, tmp_path, monkeypatch
):
    run = make_run(db, date(2026, 7, 8))
    bundle = make_daily_bundle()
    build_preview(db, run.id, "daily", bundle=bundle)
    template = _daily_template(tmp_path / "日报模板_2026-07-07.xlsx")
    original_commit = db.commit
    raised = False

    def ambiguous_commit():
        nonlocal raised
        should_raise = not raised and any(
            isinstance(obj, PublicationAttempt)
            and obj.status == "metadata_committed"
            for obj in db.identity_map.values()
        )
        original_commit()
        if should_raise:
            raised = True
            raise ConnectionError("commit acknowledgement lost")

    monkeypatch.setattr(db, "commit", ambiguous_commit)

    with pytest.raises(PublicationRecoveryRequired):
        publish(
            db,
            run.id,
            ["daily"],
            "local-operator",
            bundles={"daily": bundle},
            template_paths={"daily": template},
            output_dir=tmp_path,
        )

    attempt = db.scalar(
        select(PublicationAttempt).where(PublicationAttempt.run_id == run.id)
    )
    assert attempt.status == "metadata_committed"
    assert Path(attempt.staging_path).is_dir()
    assert publication_repo.current_daily(db, run.report_date) is not None

    assert recover_publication_attempts(db, tmp_path) == [attempt.id]
    assert Path(attempt.final_path).is_dir()
