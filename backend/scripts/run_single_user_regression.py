"""Single-user release regression with redacted, repository-external output."""

from __future__ import annotations

import argparse
import logging
import hashlib
import json
import tempfile
from copy import copy
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

import app.models  # noqa: F401  Register all ORM tables.
from app.core import constants as C
from app.core.database import Base
from app.domain.fact_bundle import FactBundle
from app.models.publication import PublishedReport, ReportArtifact
from app.models.runs import ReportRun, RunStatus
from app.pipeline.calculation import validators, weekly as weekly_calc
from app.pipeline.calculation.daily import ITEMS
from app.pipeline.cleansing import cleanse
from app.pipeline.export import weekly_exporter
from app.pipeline.input import parsers
from app.pipeline.input.daily_workbook import parse_daily_workbook
from app.repositories import input_repo, report_repo, run_repo
from app.services import publication_service
from app.services.preview_service import build_preview
from scripts import run_chain_regression


REPO_ROOT = Path(__file__).resolve().parents[1]
RELEASE_CHECKS = (
    "daily_values",
    "daily_date_columns",
    "daily_styles",
    "daily_merges",
    "tenure_controls",
    "weekly_values",
    "weekly_styles",
    "weekly_merges",
    "daily_weekly_reconciliation",
    "manifest",
    "event_ledger",
    "validation_report",
)


def _suppress_pipeline_logs(function):
    @wraps(function)
    def wrapped(*args, **kwargs):
        previous = logging.root.manager.disable
        logging.disable(logging.CRITICAL)
        try:
            return function(*args, **kwargs)
        finally:
            logging.disable(previous)

    return wrapped


def _require_external(path: Path, label: str) -> Path:
    resolved = path.expanduser().resolve()
    if resolved == REPO_ROOT or resolved.is_relative_to(REPO_ROOT):
        raise ValueError(f"{label} must be outside repository")
    return resolved


def _dated_directories(input_root: Path) -> list[date]:
    dates: list[date] = []
    for candidate in input_root.iterdir():
        if not candidate.is_dir():
            continue
        try:
            report_date = date.fromisoformat(candidate.name)
        except ValueError:
            continue
        try:
            run_chain_regression.discover_finalized_workbook(
                input_root, report_date
            )
        except RuntimeError:
            continue
        dates.append(report_date)
    return sorted(dates)


def _date_headers(path: Path) -> list[str]:
    sheet = load_workbook(path, read_only=True, data_only=False)["Sheet1"]
    headers: list[str] = []
    for cell in sheet[1]:
        value = cell.value
        if isinstance(value, datetime):
            headers.append(value.date().isoformat())
        elif isinstance(value, date):
            headers.append(value.isoformat())
    return headers


def _daily_case_summary(
    chain_case: dict[str, Any], generated: Path, expected: Path
) -> dict[str, Any]:
    workbook = compare_workbooks_redacted(generated, expected)
    checks = {
        "calculation": bool(chain_case.get("passed")),
        "values": not workbook["value_mismatches"],
        "date_columns": _date_headers(generated) == _date_headers(expected),
        "styles": not workbook["style_mismatches"],
        "merges": not workbook["merge_mismatches"],
        "tenure": bool(chain_case.get("tenure_total_match"))
        and not chain_case.get("tenure_row_mismatches"),
    }
    return {
        "report_date": chain_case["report_date"],
        "status": "passed" if all(checks.values()) else "failed",
        "checks": checks,
        "row_mismatches": list(chain_case.get("row_mismatches") or ()),
        "tenure_row_mismatches": list(
            chain_case.get("tenure_row_mismatches") or ()
        ),
        "hard_failures": list(chain_case.get("hard_failures") or ()),
        "workbook_mismatches": {
            key: value
            for key, value in workbook.items()
            if key.endswith("_mismatches") and value
        },
    }


def _write_redacted_summary(output_root: Path, summary: dict[str, Any]) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "regression_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _finalized_weekly(input_root: Path, week_end: date) -> Path | None:
    folder = input_root / week_end.isoformat()
    if not folder.is_dir():
        return None
    compact = week_end.strftime("%Y%m%d")
    matches = sorted(
        path
        for path in folder.iterdir()
        if path.is_file()
        and path.name.startswith("员工数增减周报")
        and compact in path.name
        and path.suffix.lower() == ".xlsx"
    )
    return matches[0] if len(matches) == 1 else None


def _run_weekly_case(input_root: Path, week_end: date, temp_output: Path) -> dict[str, Any]:
    expected = _finalized_weekly(input_root, week_end)
    if expected is None:
        raise RuntimeError("finalized weekly workbook missing")
    week_start = week_end - timedelta(days=week_end.weekday())
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    try:
        with Session() as db:
            current = week_start
            while current <= week_end:
                folder = input_root / current.isoformat()
                compact = current.strftime("%Y%m%d")
                employees_path = run_chain_regression._find_one(
                    folder, "人员表", {".xls", ".xlsx"}, compact
                )
                resignations_path = run_chain_regression._find_one(
                    folder, "离职人员报表", {".xls", ".xlsx"}, compact
                )
                employees = cleanse.normalize_employee_types(
                    parsers.parse_employees(str(employees_path))
                )
                employees, _ = cleanse.filter_inclusion(employees)
                employees = cleanse.trace_resign_transfer(employees)
                input_repo.upsert_employees(db, employees, current)
                input_repo.upsert_resignations(
                    db, parsers.parse_resignations(str(resignations_path))
                )
                finalized_daily = run_chain_regression.discover_finalized_workbook(
                    input_root, current
                )
                rows, _ = parse_daily_workbook(finalized_daily, current)
                report_repo.save_daily(db, current, rows, commit=False)
                db.commit()
                current += timedelta(days=1)

            context = weekly_calc.compute_weekly(db, week_start, week_end)
            context["validations"] = validators.run_weekly_checks(context)
            generated = Path(
                weekly_exporter.export_weekly(context, str(temp_output))
            )
            workbook = compare_workbooks_redacted(generated, expected)
            hard_failures = [
                str(item.get("check") or "unnamed_hard_check")
                for item in validators.hard_failures(context["validations"])
            ]
            checks = {
                "values": not workbook["value_mismatches"],
                "styles": not workbook["style_mismatches"],
                "merges": not workbook["merge_mismatches"],
                "daily_weekly_reconciliation": not hard_failures
                and bool((context.get("daily_reconciliation") or {}).get("complete")),
            }
            return {
                "week_start": week_start.isoformat(),
                "week_end": week_end.isoformat(),
                "status": "passed" if all(checks.values()) else "failed",
                "checks": checks,
                "hard_failures": hard_failures,
                "workbook_mismatches": {
                    key: value
                    for key, value in workbook.items()
                    if key.endswith("_mismatches") and value
                },
            }
    finally:
        engine.dispose()


@_suppress_pipeline_logs
def run_regression(
    input_root: Path,
    output_root: Path,
    *,
    start: date = date(2026, 7, 8),
    end: date | None = None,
    structured_image_dir: Path | None = None,
) -> dict[str, Any]:
    """Replay accepted mentor-local dates; persist only redacted evidence."""
    source = _require_external(input_root, "input_root")
    output = _require_external(output_root, "output_root")
    if output == source or output.is_relative_to(source):
        raise ValueError("output_root must be outside input_root")
    if structured_image_dir is not None:
        structured_image_dir = _require_external(
            structured_image_dir, "structured_image_dir"
        )

    summary: dict[str, Any] = {
        "mode": "mentor-local",
        "status": "blocked",
        "start": start.isoformat(),
        "end": end.isoformat() if end else None,
        "baseline_date": None,
        "cases": [],
        "weekly_cases": [],
    }
    if not source.is_dir():
        summary["error_code"] = "input_root_missing"
        _write_redacted_summary(output, summary)
        return summary

    available = _dated_directories(source)
    candidates = [day for day in available if day >= start and (end is None or day <= end)]
    baselines = [day for day in available if day < start]
    if not candidates or not baselines:
        summary["error_code"] = "accepted_date_or_baseline_missing"
        _write_redacted_summary(output, summary)
        return summary
    baseline_date = max(baselines)
    summary["baseline_date"] = baseline_date.isoformat()
    summary["end"] = candidates[-1].isoformat()

    with tempfile.TemporaryDirectory(prefix="hr_single_user_regression_") as temp:
        for index, report_date in enumerate(candidates):
            temp_output = Path(temp) / "daily" / report_date.isoformat()
            args = SimpleNamespace(
                data_root=source,
                baseline_date=baseline_date,
                dates=candidates[: index + 1],
                structured_image_dir=structured_image_dir,
                output_dir=temp_output,
            )
            try:
                chain_summary = run_chain_regression.run(args)
                summary["baseline_tenure_match"] = bool(
                    chain_summary.get("baseline_tenure_match")
                )
                chain_case = next(
                    case
                    for case in chain_summary.get("cases") or ()
                    if case.get("report_date") == report_date.isoformat()
                )
                if not chain_case.get("passed"):
                    case_summary = {
                        "report_date": chain_case["report_date"],
                        "status": "failed",
                        "checks": {"calculation": False},
                        "row_mismatches": list(
                            chain_case.get("row_mismatches") or ()
                        ),
                        "tenure_row_mismatches": list(
                            chain_case.get("tenure_row_mismatches") or ()
                        ),
                        "hard_failures": list(
                            chain_case.get("hard_failures") or ()
                        ),
                        "workbook_mismatches": {},
                    }
                    summary["cases"].append(case_summary)
                    summary["status"] = "failed"
                    summary["error_code"] = "daily_regression_failed"
                    break
                generated = temp_output / (
                    f"员工数增减情况日报_{report_date.isoformat()}.xlsx"
                )
                if not generated.is_file():
                    raise FileNotFoundError("generated daily workbook missing")
                expected = run_chain_regression.discover_finalized_workbook(
                    source, report_date
                )
                case_summary = _daily_case_summary(
                    chain_case, generated, expected
                )
                summary["cases"].append(case_summary)
                if case_summary["status"] != "passed":
                    summary["status"] = "failed"
                    summary["error_code"] = "daily_regression_failed"
                    break
            except Exception as exc:
                error_code = f"pipeline_{type(exc).__name__}"
                summary["cases"].append(
                    {
                        "report_date": report_date.isoformat(),
                        "status": "blocked",
                        "error_code": error_code,
                    }
                )
                summary["status"] = "blocked"
                summary["error_code"] = error_code
                break
        else:
            passed = (
                summary.get("baseline_tenure_match") is True
                and len(summary["cases"]) == len(candidates)
                and all(case["status"] == "passed" for case in summary["cases"])
            )
            summary["status"] = "passed" if passed else "failed"
            if not passed:
                summary["error_code"] = "daily_regression_failed"

        for week_end in candidates:
            if _finalized_weekly(source, week_end) is None:
                continue
            try:
                summary["weekly_cases"].append(
                    _run_weekly_case(source, week_end, Path(temp) / "weekly")
                )
            except Exception:
                summary["weekly_cases"].append(
                    {
                        "week_end": week_end.isoformat(),
                        "status": "blocked",
                        "error_code": "weekly_input_or_calculation_failed",
                    }
                )
        if summary["status"] == "passed" and any(
            case["status"] != "passed" for case in summary["weekly_cases"]
        ):
            summary["status"] = "failed"
            summary["error_code"] = "weekly_regression_failed"

    _write_redacted_summary(output, summary)
    return summary


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _fake_employments(report_date: date) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "person_key": "fake-person-1",
                "person_id": "fake-person-1",
                "emp_no": "FAKE-E1",
                "employee_type": "正式员工",
                "employee_status": "active",
                "hire_date": report_date,
                "leave_date": None,
                "hire_first_visible": report_date,
                "leave_first_visible": None,
                "business_unit": "NENT",
                "business_unit_no": "NENT",
                "project_no": "FAKE-P1",
                "project_name": "测试项目",
            }
        ]
    )


def _fake_bundle(report_date: date, *, weekly: bool = False) -> FactBundle:
    expected_dates = [
        "2026-07-06",
        "2026-07-07",
        "2026-07-08",
        "2026-07-09",
        "2026-07-10",
    ]
    return FactBundle(
        report_date=report_date,
        baseline_date=date(2026, 7, 7),
        rule_version="fake-regression-v1",
        employments=_fake_employments(date(2026, 7, 8)),
        events=pd.DataFrame(
            [
                {
                    "event_key": "fake-hire-1",
                    "event_type": "hire",
                    "source_type": "personnel",
                    "source_event_ref": "fake-row-1",
                    "effective_date": date(2026, 7, 8),
                    "first_visible_date": date(2026, 7, 8),
                    "classification": "deterministic",
                }
            ]
        ),
        baseline_rows={8: 0, 9: 0, 13: 0, 14: 0, 30: 0},
        daily_reconciliation=(
            {
                "available_days": 5,
                "report_dates": expected_dates,
                "expected_dates": expected_dates,
                "complete": True,
                "joiners": 1,
                "leavers": 0,
            }
            if weekly
            else {}
        ),
    )


def _fake_daily_template(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "事项")
    sheet.cell(1, 2, datetime(2026, 7, 7))
    sheet.cell(1, 2).number_format = "yyyy-mm-dd"
    for row in range(2, 41):
        sheet.cell(row, 1, ITEMS.get(row, f"保留行{row}"))
        sheet.cell(row, 2, 0)
        sheet.cell(row, 2).number_format = "0"
    for row in (25, 29, 36):
        sheet.cell(row, 2, datetime(2026, 7, 7))
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"
    sheet.merge_cells("A42:B42")
    sheet.cell(42, 1, "fake-layout-anchor")

    tenure = workbook.create_sheet("在岗时长")
    tenure.append(["事业部", "YTD离职人数", "平均在职（年）", datetime(2026, 7, 7)])
    for label in C.get_tenure_bu_labels():
        tenure.append([label, 0, None, None])
    tenure.append(["合计", 0, None, None])
    workbook.save(path)
    return path


def _new_run(db, report_date: date) -> ReportRun:
    baseline = db.scalar(
        select(PublishedReport)
        .where(
            PublishedReport.report_kind == "daily",
            PublishedReport.period_end < report_date,
            PublishedReport.is_current.is_(True),
            PublishedReport.is_deleted == 0,
        )
        .order_by(PublishedReport.period_end.desc(), PublishedReport.version.desc())
        .limit(1)
    )
    run = ReportRun(
        report_date=report_date,
        status=RunStatus.ready.value,
        rule_version="fake-regression-v1",
        baseline_report_id=baseline.id if baseline else None,
    )
    db.add(run)
    db.flush()
    run_repo.ensure_report_targets(db, run.id)
    db.commit()
    return run


def _artifact_paths(db, report_id: str) -> dict[str, Path]:
    artifacts = db.scalars(
        select(ReportArtifact).where(ReportArtifact.report_id == report_id)
    ).all()
    return {artifact.artifact_kind: Path(artifact.protected_path) for artifact in artifacts}


def _manifest_matches(paths: dict[str, Path]) -> bool:
    manifest = json.loads(paths["manifest"].read_text(encoding="utf-8"))
    listed = manifest.get("artifacts") or {}
    expected = set(paths) - {"manifest"}
    return set(listed) == expected and all(
        listed[kind].get("sha256") == _sha256(paths[kind])
        and listed[kind].get("size_bytes") == paths[kind].stat().st_size
        for kind in expected
    )


def compare_workbooks_redacted(
    actual_path: Path, expected_path: Path, *, mismatch_limit: int = 200
) -> dict[str, Any]:
    """Compare workbook content/layout while returning coordinates only."""
    actual = load_workbook(actual_path, data_only=False)
    expected = load_workbook(expected_path, data_only=False)
    result: dict[str, Any] = {
        "sheet_mismatches": [],
        "dimension_mismatches": [],
        "value_mismatches": [],
        "style_mismatches": [],
        "merge_mismatches": [],
    }

    def same_style(left, right) -> bool:
        return (
            copy(left.font) == copy(right.font)
            and copy(left.fill) == copy(right.fill)
            and copy(left.border) == copy(right.border)
            and copy(left.alignment) == copy(right.alignment)
            and left.number_format == right.number_format
            and copy(left.protection) == copy(right.protection)
        )

    if actual.sheetnames != expected.sheetnames:
        result["sheet_mismatches"].append("workbook")

    for sheet_name in expected.sheetnames:
        if sheet_name not in actual.sheetnames:
            result["sheet_mismatches"].append(sheet_name)
            continue
        actual_sheet = actual[sheet_name]
        expected_sheet = expected[sheet_name]
        if (actual_sheet.max_row, actual_sheet.max_column) != (
            expected_sheet.max_row,
            expected_sheet.max_column,
        ):
            result["dimension_mismatches"].append(sheet_name)
        max_row = max(actual_sheet.max_row, expected_sheet.max_row)
        max_column = max(actual_sheet.max_column, expected_sheet.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                actual_cell = actual_sheet.cell(row, column)
                expected_cell = expected_sheet.cell(row, column)
                coordinate = f"{sheet_name}!{expected_cell.coordinate}"
                if (
                    actual_cell.value != expected_cell.value
                    and len(result["value_mismatches"]) < mismatch_limit
                ):
                    result["value_mismatches"].append(coordinate)
                if (
                    not same_style(actual_cell, expected_cell)
                    and len(result["style_mismatches"]) < mismatch_limit
                ):
                    result["style_mismatches"].append(coordinate)
        if {
            str(value) for value in actual_sheet.merged_cells.ranges
        } != {str(value) for value in expected_sheet.merged_cells.ranges}:
            result["merge_mismatches"].append(sheet_name)

    result["passed"] = not any(result[key] for key in result)
    return result


@_suppress_pipeline_logs
def run_fake_regression(output_root: Path) -> dict[str, Any]:
    """Publish fake daily/weekly reports and verify every release artifact."""
    output = _require_external(output_root, "output_root")
    output.mkdir(parents=True, exist_ok=True)
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    checks = {check: False for check in RELEASE_CHECKS}

    with Session() as db:
        daily_run = _new_run(db, date(2026, 7, 8))
        daily_bundle = _fake_bundle(daily_run.report_date)
        daily_preview = build_preview(
            db, daily_run.id, "daily", bundle=daily_bundle
        )
        template = _fake_daily_template(output / "fake-baseline.xlsx")
        daily_report = publication_service.publish(
            db,
            daily_run.id,
            ["daily"],
            "fake-regression",
            bundles={"daily": daily_bundle},
            template_paths={"daily": str(template)},
            output_dir=output,
        )[0]
        daily_paths = _artifact_paths(db, daily_report.id)
        daily_projection = publication_service._daily_excel_projection(
            daily_preview, daily_paths["excel"]
        )
        daily_book = load_workbook(daily_paths["excel"])
        daily_sheet = daily_book["Sheet1"]
        checks["daily_values"] = (
            daily_projection["expected_rows"] == daily_projection["actual_rows"]
        )
        checks["daily_date_columns"] = (
            daily_projection["file_date"] == "2026-07-08"
            and daily_sheet.cell(1, daily_sheet.max_column).value.date()
            == date(2026, 7, 8)
        )
        checks["daily_styles"] = all(
            daily_sheet.cell(row, 2)._style == daily_sheet.cell(row, 3)._style
            for row in range(1, 41)
        )
        checks["daily_merges"] = "A42:B42" in {
            str(value) for value in daily_sheet.merged_cells.ranges
        }
        tenure_rows = daily_projection["actual_tenure"]
        checks["tenure_controls"] = (
            daily_projection["expected_tenure"] == tenure_rows
            and sum(int(row.get("ytd_leavers") or 0) for row in tenure_rows)
            == int(daily_preview.rows[14].value or 0)
            and daily_book["在岗时长"].cell(1, 4).value.date()
            == date(2026, 7, 8)
        )

        weekly_run = _new_run(db, date(2026, 7, 10))
        weekly_bundle = _fake_bundle(weekly_run.report_date, weekly=True)
        weekly_preview = build_preview(
            db,
            weekly_run.id,
            "weekly",
            bundle=weekly_bundle,
            week_start=date(2026, 7, 6),
            week_end=date(2026, 7, 10),
        )
        weekly_report = publication_service.publish(
            db,
            weekly_run.id,
            ["weekly"],
            "fake-regression",
            bundles={"weekly": weekly_bundle},
            periods={"weekly": (date(2026, 7, 6), date(2026, 7, 10))},
            output_dir=output,
        )[0]
        weekly_paths = _artifact_paths(db, weekly_report.id)
        actual_main, actual_cc = publication_service._parse_weekly_excel(
            weekly_paths["excel"]
        )
        checks["weekly_values"] = (
            publication_service._weekly_main_projection(weekly_preview.main_rows)
            == publication_service._weekly_main_projection(actual_main)
            and publication_service._weekly_cc_projection(weekly_preview.cc_rows)
            == publication_service._weekly_cc_projection(actual_cc)
        )
        weekly_book = load_workbook(weekly_paths["excel"])
        sheet2 = weekly_book["Sheet2"]
        checks["weekly_styles"] = all(
            sheet2.cell(1, column).font.bold
            for column in (1, 2, 3, 4, 7, 8, 9, 10)
        )
        required_merges = {"A1:A2", "B1:B2", "C1:C2", "D1:F1", "G1:G2", "H1:H2", "I1:I2", "J1:J2"}
        checks["weekly_merges"] = required_merges.issubset(
            {str(value) for value in sheet2.merged_cells.ranges}
        )
        weekly_joiners = sum(
            int(row.get("joiners") or 0) for row in weekly_preview.main_rows
        )
        checks["daily_weekly_reconciliation"] = (
            int(daily_preview.rows[2].value or 0)
            == int(weekly_bundle.daily_reconciliation["joiners"])
            == weekly_joiners
        )

        all_paths = (daily_paths, weekly_paths)
        checks["manifest"] = all(_manifest_matches(paths) for paths in all_paths)
        checks["event_ledger"] = all(
            isinstance(json.loads(paths["event_ledger"].read_text(encoding="utf-8")), list)
            for paths in all_paths
        )
        checks["validation_report"] = all(
            json.loads(paths["validation_report"].read_text(encoding="utf-8")).get(
                "publishable"
            )
            is True
            for paths in all_paths
        )

    engine.dispose()
    return {
        "mode": "fake",
        "status": "passed" if all(checks.values()) else "failed",
        "checks": checks,
    }


def _cli_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("date must use YYYY-MM-DD") from exc


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run fake or mentor-local HR report release regression."
    )
    parser.add_argument("--mode", choices=("fake", "mentor-local"), required=True)
    parser.add_argument("--input-root", type=Path)
    parser.add_argument("--output-root", type=Path)
    parser.add_argument("--structured-image-dir", type=Path)
    parser.add_argument("--start", type=_cli_date, default=date(2026, 7, 8))
    parser.add_argument("--end", type=_cli_date)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if args.mode == "fake":
        if args.output_root is not None:
            summary = run_fake_regression(args.output_root)
        else:
            with tempfile.TemporaryDirectory(prefix="hr_fake_regression_") as temp:
                summary = run_fake_regression(Path(temp))
        passed = sum(bool(value) for value in summary["checks"].values())
        print(f"fake {'PASS' if summary['status'] == 'passed' else 'FAIL'} checks={passed}/{len(RELEASE_CHECKS)}")
        return 0 if summary["status"] == "passed" else 1

    if args.input_root is None or args.output_root is None:
        _parser().error("mentor-local mode requires --input-root and --output-root")
    try:
        summary = run_regression(
            args.input_root,
            args.output_root,
            start=args.start,
            end=args.end,
            structured_image_dir=args.structured_image_dir,
        )
    except ValueError:
        print("mentor-local BLOCKED configuration_error")
        return 2
    daily_passed = sum(case.get("status") == "passed" for case in summary["cases"])
    weekly_passed = sum(
        case.get("status") == "passed" for case in summary["weekly_cases"]
    )
    print(
        f"mentor-local {summary['status'].upper()} "
        f"daily={daily_passed}/{len(summary['cases'])} "
        f"weekly={weekly_passed}/{len(summary['weekly_cases'])}"
    )
    return 0 if summary["status"] == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
