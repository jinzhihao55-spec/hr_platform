"""openpyxl 共享样式。"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")  # 绿底=派生公式行
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 已验收周报模板的固定视觉契约。
WEEKLY_HEADER_BLUE = PatternFill("solid", fgColor="DDEBF7")
WEEKLY_HEADER_GREEN = PatternFill("solid", fgColor="E2F0D9")
WEEKLY_HEADER_PEACH = PatternFill("solid", fgColor="FCE4D6")
WEEKLY_HEADER_YELLOW = PatternFill("solid", fgColor="FFFF00")
WEEKLY_HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True)
WEEKLY_BODY_FONT = Font(name="Microsoft YaHei", size=10)
WEEKLY_TOTAL_FONT = Font(
    name="Microsoft YaHei", size=10, bold=True, color="FF0000"
)
WEEKLY_THIN = Side(style="thin", color="D9D9D9")
WEEKLY_BORDER = Border(
    left=WEEKLY_THIN,
    right=WEEKLY_THIN,
    top=WEEKLY_THIN,
    bottom=WEEKLY_THIN,
)
