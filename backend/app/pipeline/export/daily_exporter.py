"""日报 Excel 导出（纯代码，§2.3），对齐人工定稿格式：
工作簿 = Sheet1 主表(Row2-40，按日追加日期列) + 在岗时长（每日原地刷新）。

定稿事实（template_mapping.md + project/日报 实测）：
- 月内逐日在 Sheet1 追加一列（表头为 datetime、继承前一列样式），历史列不改动；
- 月初文件从 HR 已确认的 month-opening 模板派生：A 列标签/样式承袭，
  B 列使用独立月初基线值；不得覆盖上月 daily_reports 或静默猜测跨月重述；
- Row25/29/36 是分节日期戳，各列值恒等于该列报告日；
- 在岗时长 sheet 固定 10×4，报告日只写在 D1，原地刷新值、不重建（保留定稿样式）；
- 找不到任何可承袭的工作簿时必须阻断（DailyTemplateMissingError），
  禁止凭空造模板——首次部署应先导入最近一份定稿。"""
from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.core.exceptions import DailyTemplateMissingError
from app.pipeline.export import styles as S

# 分节日期戳行：值 = 该列报告日（与定稿一致）
_SECTION_DATE_ROWS = (25, 29, 36)

_DAILY_NAME_RE = re.compile(r"员工数增减情况日报_(\d{4}-\d{2}-\d{2})\.xlsx$")


def find_previous_daily_workbook(out_dir: str, report_date: date) -> str | None:
    """查找可承袭的前一份日报工作簿（允许跨月——月初据此派生新文件）。

    优先 `{out_dir}/finalized/`（导入定稿时留存的副本），其次 out_dir 本身
    （此前生成的输出）。同日期时定稿副本优先。"""
    candidates: list[tuple[date, int, Path]] = []
    for priority, folder in enumerate((Path(out_dir) / "finalized", Path(out_dir))):
        if not folder.is_dir():
            continue
        for path in folder.glob("员工数增减情况日报_*.xlsx"):
            matched = _DAILY_NAME_RE.search(path.name)
            if not matched:
                continue
            day = date.fromisoformat(matched.group(1))
            if day < report_date:
                candidates.append((day, priority, path))
    if not candidates:
        return None
    best_day = max(day for day, _, _ in candidates)
    best = min(
        (item for item in candidates if item[0] == best_day),
        key=lambda item: item[1],
    )
    return str(best[2])


def export_daily(
    ctx: dict,
    tenure: dict,
    out_dir: str,
    baseline_values: dict[int, float] | None = None,
    template_path: str | None = None,
) -> str:
    report_date: date = ctx["report_date"]
    rows = ctx["rows"]

    if not (template_path and Path(template_path).is_file()):
        raise DailyTemplateMissingError(
            f"缺少可承袭的日报工作簿，无法生成 {report_date} 日报：请先通过 "
            "POST /reports/daily/import 导入最近一份已验收定稿（禁止凭空造模板）。",
            detail={"report_date": report_date.isoformat()},
        )

    wb = load_workbook(template_path)
    ws = wb["Sheet1"]
    last_header = ws.cell(1, ws.max_column).value
    same_month = (
        isinstance(last_header, (date, datetime))
        and (last_header.year, last_header.month) == (report_date.year, report_date.month)
    )
    if not same_month:
        _collapse_to_month_baseline(ws, baseline_values or {})
    _append_report_column(ws, report_date, rows)
    _refresh_tenure_sheet(wb, tenure, report_date)

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    path = str(Path(out_dir) / f"员工数增减情况日报_{report_date.isoformat()}.xlsx")
    wb.save(path)
    return path


def _cell_value(rows: dict, row_no: int, report_date: date):
    if row_no in _SECTION_DATE_ROWS:
        return datetime(report_date.year, report_date.month, report_date.day)
    info = rows.get(row_no) or {}
    if info.get("is_blank"):
        return None
    return info.get("value")


def _collapse_to_month_baseline(ws, baseline_values: dict[int, float]) -> None:
    """月初派生：只保留 A 标签列 + 上月末列（成为 B 基线列），
    并以 HR 已确认的独立 month-opening 行值覆盖 B 列。"""
    if ws.max_column > 2:
        last_letter = ws.cell(1, ws.max_column).column_letter
        last_width = ws.column_dimensions[last_letter].width
        ws.delete_cols(2, ws.max_column - 2)
        ws.column_dimensions["B"].width = last_width
    for row_no, value in baseline_values.items():
        ws.cell(int(row_no), 2).value = value


def _append_report_column(ws, report_date: date, rows: dict) -> None:
    """续列：历史列原样保留，追加报告日一列并逐 cell 继承前一列样式。"""
    prev_col = ws.max_column
    new_col = prev_col + 1
    for row_no in range(1, 41):
        source = ws.cell(row_no, prev_col)
        target = ws.cell(row_no, new_col)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        if row_no == 1:
            target.value = datetime(report_date.year, report_date.month, report_date.day)
        else:
            target.value = _cell_value(rows, row_no, report_date)
    prev_letter = ws.cell(1, prev_col).column_letter
    ws.column_dimensions[ws.cell(1, new_col).column_letter].width = (
        ws.column_dimensions[prev_letter].width
    )


def _refresh_tenure_sheet(wb: Workbook, tenure: dict, report_date: date) -> None:
    """在岗时长 sheet：原地刷新值（保留定稿样式），报告日仅写 D1。

    模板缺该 sheet 时才新建（正常链路不会走到——模板都来自定稿/生成物）。"""
    if "在岗时长" not in wb.sheetnames:
        _create_tenure_sheet(wb)
    ts = wb["在岗时长"]
    ts.cell(1, 4).value = datetime(report_date.year, report_date.month, report_date.day)
    line = 2
    for r in tenure["rows"]:
        ts.cell(line, 1).value = r.get("business_unit") or r.get("slot")
        ts.cell(line, 2).value = r["ytd_leavers"]
        ts.cell(line, 3).value = r["avg_tenure_years"]
        line += 1
    t = tenure["total"]
    ts.cell(line, 1).value = t["business_unit"]
    ts.cell(line, 2).value = t["ytd_leavers"]
    ts.cell(line, 3).value = t["avg_tenure_years"]


def _create_tenure_sheet(wb: Workbook) -> None:
    ts = wb.create_sheet("在岗时长")
    ts.append(["事业部", "YTD离职人数", "平均在职（年）", None])
    for c in ts[1]:
        c.font = S.BOLD
        c.fill = S.HEADER_FILL
        c.alignment = S.CENTER
        c.border = S.BORDER
    ts.cell(1, 4).number_format = "yyyy-mm-dd"
    for _ in range(9):
        ts.append([None, None, None, None])
    for col in "ABCD":
        ts.column_dimensions[col].width = 16
