"""周报 Excel 导出（纯代码，§2.4）。"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.core import constants as C
from app.pipeline.export import styles as S


def _style_cell(cell, *, font, fill=None, alignment=S.CENTER) -> None:
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = S.WEEKLY_BORDER


def export_weekly(ctx: dict, out_dir: str) -> str:
    week_end: date = ctx["week_end"]
    wb = Workbook()

    # ---- Sheet2 主体 × 事业部 ----
    ws = wb.active
    ws.title = "Sheet2"
    # 第1/2行表头（员工类型跨3个子列）
    headers = ["主体", "事业部", "在职总数", "在职员工类型", None, None,
               "本周入职人员", "本周离职人员", "部门前三大项目名称", "项目人数"]
    ws.append(headers)
    ws.append([None, None, None, "正式员工", "实习生", "劳务人员", None, None, None, None])
    header_fills = {
        1: S.WEEKLY_HEADER_BLUE,
        2: S.WEEKLY_HEADER_BLUE,
        3: S.WEEKLY_HEADER_BLUE,
        4: S.WEEKLY_HEADER_GREEN,
        7: S.WEEKLY_HEADER_PEACH,
        8: S.WEEKLY_HEADER_PEACH,
        9: S.WEEKLY_HEADER_YELLOW,
        10: S.WEEKLY_HEADER_YELLOW,
    }
    for column, fill in header_fills.items():
        _style_cell(
            ws.cell(1, column), font=S.WEEKLY_HEADER_FONT, fill=fill
        )
    for column in range(4, 7):
        _style_cell(
            ws.cell(2, column),
            font=S.WEEKLY_HEADER_FONT,
            fill=S.WEEKLY_HEADER_GREEN,
        )
    ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2"); ws.merge_cells("C1:C2")
    ws.merge_cells("D1:F1")
    ws.merge_cells("G1:G2"); ws.merge_cells("H1:H2")
    ws.merge_cells("I1:I2"); ws.merge_cells("J1:J2")

    first_data_row = ws.max_row + 1
    top3_total = 0
    for r in ctx["main_rows"]:
        projects = r.get("top3_projects") or [{"name": None, "count": None}]
        block_start = ws.max_row + 1
        for index, project in enumerate(projects):
            count = project.get("count")
            if count is not None:
                top3_total += int(count)
            ws.append([
                None,
                r["business_unit"] if index == 0 else None,
                r["headcount"] if index == 0 else None,
                r["cnt_formal"] if index == 0 else None,
                r["cnt_intern"] if index == 0 else None,
                r["cnt_labor"] if index == 0 else None,
                r["joiners"] if index == 0 else None,
                r["leavers"] if index == 0 else None,
                project.get("name"),
                count,
            ])
            for column in range(1, 11):
                _style_cell(
                    ws.cell(ws.max_row, column),
                    font=S.WEEKLY_BODY_FONT,
                    alignment=S.LEFT if column == 9 else S.CENTER,
                )
        block_end = ws.max_row
        if block_end > block_start:
            for column in "BCDEFGH":
                ws.merge_cells(f"{column}{block_start}:{column}{block_end}")

    last_data_row = ws.max_row
    if last_data_row >= first_data_row:
        ws.cell(first_data_row, 1, C.WEEKLY_SUBJECT_LABEL)
        if last_data_row > first_data_row:
            ws.merge_cells(
                start_row=first_data_row,
                start_column=1,
                end_row=last_data_row,
                end_column=1,
            )
    # 合计
    mr = ctx["main_rows"]
    if mr:
        ws.append(["合计", None, sum(x["headcount"] for x in mr),
                   sum(x["cnt_formal"] for x in mr), sum(x["cnt_intern"] for x in mr),
                   sum(x["cnt_labor"] for x in mr), sum(x["joiners"] for x in mr),
                   sum(x["leavers"] for x in mr), None, top3_total])
        for cell in ws[ws.max_row]:
            _style_cell(cell, font=S.WEEKLY_TOTAL_FONT)
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24
    for index, width in enumerate((8, 12, 12, 12, 12, 12, 14, 14, 38, 12), 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    # ---- Sheet1 成本中心 × 项目 ----
    s1 = wb.create_sheet("Sheet1")
    s1.append(["成本中心", "项目", "在职人数", "本周入职人员", "本周离职人员"])
    for c in s1[1]:
        _style_cell(
            c, font=S.WEEKLY_HEADER_FONT, fill=S.WEEKLY_HEADER_BLUE
        )
    for r in ctx["cc_rows"]:
        s1.append([r["cost_center"], r["project"], r["headcount"],
                   r["joiners"] or None, r["leavers"] or None])
        for column in range(1, 6):
            _style_cell(
                s1.cell(s1.max_row, column),
                font=S.WEEKLY_BODY_FONT,
                alignment=S.LEFT if column == 2 else S.CENTER,
            )
    for row in range(1, s1.max_row + 1):
        s1.row_dimensions[row].height = 24
    for column, width in zip("ABCDE", (12, 34, 12, 14, 14), strict=True):
        s1.column_dimensions[column].width = width

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    path = str(Path(out_dir) / f"员工数增减周报_{week_end.isoformat()}.xlsx")
    wb.save(path)
    return path
