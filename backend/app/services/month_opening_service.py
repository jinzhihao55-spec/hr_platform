"""HR 月初基线确认/上传流程。

月初重述是独立业务事实：不能覆盖上月末日报，也不能从上月数据静默猜测。
确认沿用时由 HR 指定基线日期；发生重述时上传 A+B 两列基线工作簿，且
“在岗时长”必须同样代表 B1 基线日期。
"""
from __future__ import annotations

import hashlib
import shutil
import tempfile
from copy import copy
from datetime import date, datetime
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.core.exceptions import (
    MonthOpeningBaselineError,
    MonthOpeningBaselineMissingError,
)
from app.pipeline.input.daily_workbook import parse_tenure_workbook
from app.pipeline.export import daily_exporter
from app.repositories import month_opening_repo, report_repo

_REQUIRED_CHAIN_ROWS = (8, 9, 13, 14, 30)


def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def _validate_report_month(report_month: date) -> date:
    if report_month.day != 1:
        raise MonthOpeningBaselineError(
            "report_month 必须为该月 1 日，例如 2026-08-01",
            detail={"report_month": report_month.isoformat()},
        )
    return report_month


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _artifact_path(report_month: date) -> Path:
    return (
        Path(settings.output_dir)
        / "month_opening"
        / f"month_opening_{report_month:%Y-%m}.xlsx"
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _extract_numeric_rows(ws, column: int) -> dict[int, int]:
    rows: dict[int, int] = {}
    for row in range(2, 41):
        value = ws.cell(row, column).value
        if isinstance(value, bool):
            rows[row] = int(value)
        elif isinstance(value, (int, float)):
            rows[row] = int(value)
    missing = [row for row in _REQUIRED_CHAIN_ROWS if row not in rows]
    if missing:
        raise MonthOpeningBaselineError(
            f"月初基线缺少必需数值行：{', '.join(f'Row{row}' for row in missing)}",
            detail={"missing_rows": missing},
        )
    return rows


def _copy_cell(source, target) -> None:
    target.value = source.value
    target._style = copy(source._style)
    if source.has_style:
        target.number_format = source.number_format
    target.hyperlink = copy(source.hyperlink)
    target.comment = copy(source.comment)


def _prepare_artifact(source: Path, baseline_column: int, destination: Path) -> None:
    wb = load_workbook(source)
    if "Sheet1" not in wb.sheetnames or "在岗时长" not in wb.sheetnames:
        raise MonthOpeningBaselineError("月初基线必须包含 Sheet1 和 在岗时长 两个 sheet")
    ws = wb["Sheet1"]
    if baseline_column != 2:
        for row in range(1, ws.max_row + 1):
            _copy_cell(ws.cell(row, baseline_column), ws.cell(row, 2))
        source_letter = ws.cell(1, baseline_column).column_letter
        ws.column_dimensions["B"].width = ws.column_dimensions[source_letter].width
    if ws.max_column > 2:
        ws.delete_cols(3, ws.max_column - 2)
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)


def _read_and_validate(
    source: Path,
    *,
    report_month: date,
    baseline_column: int,
) -> tuple[date, dict[int, int], list[dict]]:
    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:
        raise MonthOpeningBaselineError(f"无法读取月初基线工作簿：{exc}") from exc
    if "Sheet1" not in wb.sheetnames:
        raise MonthOpeningBaselineError("月初基线缺少 Sheet1")
    ws = wb["Sheet1"]
    if ws.max_column < baseline_column or ws.max_row < 40:
        raise MonthOpeningBaselineError("月初基线 Sheet1 至少需要 40 行及 A+B 两列")
    baseline_date = _as_date(ws.cell(1, baseline_column).value)
    if baseline_date is None or baseline_date >= report_month:
        raise MonthOpeningBaselineError(
            "月初基线表头日期必须早于 report_month",
            detail={"baseline_date": str(ws.cell(1, baseline_column).value)},
        )
    baseline_rows = _extract_numeric_rows(ws, baseline_column)
    try:
        tenure_rows = parse_tenure_workbook(source, baseline_date)
    except Exception as exc:
        raise MonthOpeningBaselineError(str(exc)) from exc
    tenure_total = sum(int(row["ytd_leavers"]) for row in tenure_rows)
    if tenure_total != baseline_rows[14]:
        raise MonthOpeningBaselineError(
            "月初基线在岗时长合计与 Sheet1 Row14 不一致",
            detail={"tenure_total": tenure_total, "row14": baseline_rows[14]},
        )
    return baseline_date, baseline_rows, tenure_rows


def _persist(
    db: Session,
    *,
    report_month: date,
    baseline_date: date,
    source_type: str,
    baseline_rows: dict[int, int],
    tenure_rows: list[dict],
    confirmed_by: str,
    prepared_artifact: Path,
) -> dict:
    confirmed_by = confirmed_by.strip()
    if not confirmed_by:
        raise MonthOpeningBaselineError("confirmed_by 不能为空")
    destination = _artifact_path(report_month)
    destination.parent.mkdir(parents=True, exist_ok=True)
    backup = destination.with_suffix(".previous.xlsx")
    if destination.is_file():
        shutil.copyfile(destination, backup)
    shutil.copyfile(prepared_artifact, destination)
    try:
        obj = month_opening_repo.save(
            db,
            report_month=report_month,
            baseline_date=baseline_date,
            source_type=source_type,
            baseline_rows=baseline_rows,
            tenure_rows=tenure_rows,
            template_sha256=_sha256(destination),
            confirmed_by=confirmed_by,
        )
        db.commit()
    except Exception:
        db.rollback()
        if backup.is_file():
            backup.replace(destination)
        else:
            destination.unlink(missing_ok=True)
        raise
    finally:
        backup.unlink(missing_ok=True)
    return _serialize(obj, baseline_rows, tenure_rows, destination)


def _serialize(obj, baseline_rows: dict[int, int], tenure_rows: list[dict], path: Path) -> dict:
    return {
        "status": "confirmed",
        "report_month": obj.report_month.isoformat(),
        "baseline_date": obj.baseline_date.isoformat(),
        "source_type": obj.source_type,
        "confirmed_by": obj.confirmed_by,
        "baseline_rows": baseline_rows,
        "tenure_rows": tenure_rows,
        "template_path": str(path),
        "template_sha256": obj.template_sha256,
    }


async def import_baseline(
    db: Session,
    *,
    report_month: date,
    confirmed_by: str,
    file: UploadFile,
) -> dict:
    report_month = _validate_report_month(report_month)
    with tempfile.TemporaryDirectory(prefix="hr_month_opening_") as tmp_dir:
        source = Path(tmp_dir) / (Path(file.filename or "opening.xlsx").name or "opening.xlsx")
        with source.open("wb") as output:
            shutil.copyfileobj(file.file, output)
        baseline_date, baseline_rows, tenure_rows = _read_and_validate(
            source, report_month=report_month, baseline_column=2,
        )
        prepared = Path(tmp_dir) / "prepared.xlsx"
        _prepare_artifact(source, 2, prepared)
        return _persist(
            db,
            report_month=report_month,
            baseline_date=baseline_date,
            source_type="uploaded",
            baseline_rows=baseline_rows,
            tenure_rows=tenure_rows,
            confirmed_by=confirmed_by,
            prepared_artifact=prepared,
        )


def confirm_carry_forward(
    db: Session,
    *,
    report_month: date,
    baseline_date: date,
    confirmed_by: str,
) -> dict:
    report_month = _validate_report_month(report_month)
    if baseline_date >= report_month:
        raise MonthOpeningBaselineError("baseline_date 必须早于 report_month")
    candidates = (
        Path(settings.output_dir) / "finalized" / f"员工数增减情况日报_{baseline_date}.xlsx",
        Path(settings.output_dir) / f"员工数增减情况日报_{baseline_date}.xlsx",
    )
    source = next((path for path in candidates if path.is_file()), None)
    if source is None:
        raise MonthOpeningBaselineError(
            f"找不到 {baseline_date} 已验收定稿，无法确认沿用",
            detail={"baseline_date": baseline_date.isoformat()},
        )
    wb = load_workbook(source, data_only=True)
    baseline_column = wb["Sheet1"].max_column
    detected_date, baseline_rows, tenure_rows = _read_and_validate(
        source, report_month=report_month, baseline_column=baseline_column,
    )
    if detected_date != baseline_date:
        raise MonthOpeningBaselineError(
            "指定 baseline_date 与定稿最后一列日期不一致",
            detail={"specified": baseline_date.isoformat(), "detected": detected_date.isoformat()},
        )
    with tempfile.TemporaryDirectory(prefix="hr_month_opening_") as tmp_dir:
        prepared = Path(tmp_dir) / "prepared.xlsx"
        _prepare_artifact(source, baseline_column, prepared)
        return _persist(
            db,
            report_month=report_month,
            baseline_date=baseline_date,
            source_type="carry_forward",
            baseline_rows=baseline_rows,
            tenure_rows=tenure_rows,
            confirmed_by=confirmed_by,
            prepared_artifact=prepared,
        )


def get_confirmed(db: Session, report_date: date) -> dict | None:
    report_month = _month_start(report_date)
    obj = month_opening_repo.get(db, report_month)
    if obj is None:
        return None
    baseline_rows, tenure_rows = month_opening_repo.decode(obj)
    path = _artifact_path(report_month)
    if not path.is_file() or _sha256(path) != obj.template_sha256:
        raise MonthOpeningBaselineError(
            "月初基线模板缺失或校验和不一致，请重新确认/上传",
            detail={"report_month": report_month.isoformat()},
        )
    return _serialize(obj, baseline_rows, tenure_rows, path)


def require_for_report(db: Session, report_date: date) -> dict:
    confirmed = get_confirmed(db, report_date)
    if confirmed is None:
        report_month = _month_start(report_date)
        raise MonthOpeningBaselineMissingError(
            f"{report_month:%Y-%m} 尚未由 HR 确认月初基线。"
            "请先确认沿用上月定稿，或上传包含 Sheet1 A+B 和在岗时长基线的工作簿。",
            detail={"report_month": report_month.isoformat()},
        )
    return confirmed


def prepare_generation(
    db: Session,
    report_date: date,
    explicit_baseline_date: date | None = None,
) -> dict:
    """解析当次计算/导出的基线。

    同月有历史日报时沿用正常链路；没有同月历史日报（跨月或首次运行）时，
    必须使用 HR 已确认的独立月初基线。月初在岗时长基线会持续使用，直到
    本月导入了更新的定稿快照（tenure 计算层按日期择新）。
    """
    db_baseline_date = explicit_baseline_date or report_repo.baseline_date(db, report_date)
    crosses_month = (
        db_baseline_date is None
        or (db_baseline_date.year, db_baseline_date.month)
        != (report_date.year, report_date.month)
    )
    opening = get_confirmed(db, report_date)
    previous_template = daily_exporter.find_previous_daily_workbook(
        settings.output_dir, report_date,
    )

    if crosses_month:
        opening = opening or require_for_report(db, report_date)
        opening_baseline_date = date.fromisoformat(opening["baseline_date"])
        return {
            "baseline_date": opening_baseline_date,
            "baseline_override": opening["baseline_rows"],
            "tenure_baseline": {
                "baseline_date": opening_baseline_date,
                "rows": opening["tenure_rows"],
            },
            "export_baseline_rows": opening["baseline_rows"],
            "template_path": opening["template_path"],
        }

    if previous_template is None:
        raise MonthOpeningBaselineError(
            "已存在同月日报基线，但缺少可续列的日报工作簿，请导入最近定稿后重试",
            detail={"report_date": report_date.isoformat()},
        )
    tenure_baseline = None
    if opening is not None:
        opening_baseline_date = date.fromisoformat(opening["baseline_date"])
        tenure_baseline = {
            "baseline_date": opening_baseline_date,
            "rows": opening["tenure_rows"],
        }
    return {
        "baseline_date": explicit_baseline_date,
        "baseline_override": None,
        "tenure_baseline": tenure_baseline,
        "export_baseline_rows": {},
        "template_path": previous_template,
    }
