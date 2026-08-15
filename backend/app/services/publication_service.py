"""Verified filesystem publication with database versioning and recovery state."""

from __future__ import annotations

import copy
import hashlib
import os
import shutil
import time
from dataclasses import asdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.core.logging import get_logger
from app.domain.fact_bundle import FactBundle
from app.models.facts import encode_json_text
from app.models.publication import (
    PublicationAttempt,
    PublishedReport,
    ReportArtifact,
)
from app.models.runs import ReportRun, RunReportTarget, RunSource, TargetStatus
from app.pipeline.export import daily_exporter, weekly_exporter
from app.pipeline.input.daily_workbook import (
    parse_daily_workbook,
    parse_tenure_workbook,
)
from app.repositories import publication_repo, report_repo, run_repo
from app.services.preview_service import PreviewSnapshot, build_preview


log = get_logger(__name__)


class PublicationFailed(RuntimeError):
    pass


class PublicationBlocked(PublicationFailed):
    pass


class PreviewRequired(PublicationFailed):
    pass


class PreviewChanged(PublicationFailed):
    pass


class PublicationRecoveryRequired(PublicationFailed):
    pass


_ARTIFACT_FILENAMES = {
    "execution_log": "执行说明.md",
    "event_ledger": "事件台账.json",
    "validation_report": "验证报告.json",
    "manifest": "产物清单.json",
}


def _now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _redacted_error(exc: BaseException) -> str:
    type_name = type(exc).__name__
    hint = {
        "DataError": "数据库字段或数据长度不兼容，请确认已执行最新迁移",
        "OperationalError": "数据库操作失败，请检查数据库状态",
        "OSError": "文件写入失败，请检查磁盘空间和输出目录权限",
    }.get(type_name, "内部操作失败，请查看本地系统日志")
    return f"{type_name}: {hint}"


def _default_daily_template(db: Session, report_date: date, output_dir: Path) -> str | None:
    artifact = db.scalar(
        select(ReportArtifact)
        .join(PublishedReport, ReportArtifact.report_id == PublishedReport.id)
        .where(
            PublishedReport.report_kind == "daily",
            PublishedReport.period_end < report_date,
            PublishedReport.is_current.is_(True),
            PublishedReport.is_deleted == 0,
            ReportArtifact.artifact_kind == "excel",
            ReportArtifact.is_deleted == 0,
        )
        .order_by(PublishedReport.period_end.desc())
        .limit(1)
    )
    if artifact is not None and Path(artifact.protected_path).is_file():
        return artifact.protected_path
    return daily_exporter.find_previous_daily_workbook(str(output_dir), report_date)


def _write_artifacts(
    db: Session,
    snapshot: PreviewSnapshot,
    staging_dir: Path,
    *,
    template_path: str | None,
    output_root: Path,
) -> dict[str, Path]:
    staging_dir.mkdir(parents=True, exist_ok=False)
    ctx = copy.deepcopy(dict(snapshot.calculation_context))
    if snapshot.report_kind == "daily":
        template = template_path or _default_daily_template(
            db, snapshot.period_end, output_root
        )
        excel_path = Path(
            daily_exporter.export_daily(
                ctx,
                ctx["tenure"],
                str(staging_dir),
                baseline_values=dict(snapshot.baseline_rows),
                template_path=template,
            )
        )
    else:
        excel_path = Path(weekly_exporter.export_weekly(ctx, str(staging_dir)))

    execution_path = staging_dir / _ARTIFACT_FILENAMES["execution_log"]
    execution_path.write_text(
        "\n".join(
            [
                f"# {snapshot.report_kind} 执行说明",
                "",
                f"- 期间：{snapshot.period_start} 至 {snapshot.period_end}",
                f"- 规则版本：{snapshot.rule_version}",
                f"- 预览哈希：{snapshot.snapshot_hash}",
                f"- 发布校验：{'通过' if snapshot.publishable else '未通过'}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    events_path = staging_dir / _ARTIFACT_FILENAMES["event_ledger"]
    events_path.write_text(
        encode_json_text(list(snapshot.events)), encoding="utf-8"
    )
    validation_path = staging_dir / _ARTIFACT_FILENAMES["validation_report"]
    validation_path.write_text(
        encode_json_text(asdict(snapshot.validation_summary)), encoding="utf-8"
    )
    return {
        "excel": excel_path,
        "execution_log": execution_path,
        "event_ledger": events_path,
        "validation_report": validation_path,
    }


def _daily_excel_projection(
    snapshot: PreviewSnapshot, excel_path: Path
) -> dict:
    parsed, file_date = parse_daily_workbook(excel_path, snapshot.period_end)
    expected_rows = {
        str(number): int(snapshot.rows[number].value or 0)
        for number in parsed
    }
    actual_rows = {
        str(number): int(info.get("value") or 0) for number, info in parsed.items()
    }
    tenure_rows = parse_tenure_workbook(excel_path, snapshot.period_end)
    expected_tenure = [
        {
            "slot": row.get("slot"),
            "ytd_leavers": int(row.get("ytd_leavers") or 0),
            "avg_tenure_years": row.get("avg_tenure_years"),
        }
        for row in snapshot.tenure.get("rows", [])
    ]
    actual_tenure = [
        {
            "slot": row.get("slot"),
            "ytd_leavers": int(row.get("ytd_leavers") or 0),
            "avg_tenure_years": row.get("avg_tenure_years"),
        }
        for row in tenure_rows
    ]
    return {
        "file_date": file_date.isoformat(),
        "expected_rows": expected_rows,
        "actual_rows": actual_rows,
        "expected_tenure": expected_tenure,
        "actual_tenure": actual_tenure,
    }


def _weekly_main_projection(rows: Sequence[Mapping]) -> list[dict]:
    return [
        {
            "business_unit": row.get("business_unit"),
            "headcount": int(row.get("headcount") or 0),
            "cnt_formal": int(row.get("cnt_formal") or 0),
            "cnt_intern": int(row.get("cnt_intern") or 0),
            "cnt_labor": int(row.get("cnt_labor") or 0),
            "joiners": int(row.get("joiners") or 0),
            "leavers": int(row.get("leavers") or 0),
            "top3_projects": [
                {"name": item.get("name"), "count": int(item.get("count") or 0)}
                for item in row.get("top3_projects") or []
                if item.get("name") is not None
            ],
        }
        for row in rows
    ]


def _weekly_cc_projection(rows: Sequence[Mapping]) -> list[dict]:
    return [
        {
            "cost_center": row.get("cost_center"),
            "project": row.get("project"),
            "headcount": int(row.get("headcount") or 0),
            "joiners": int(row.get("joiners") or 0),
            "leavers": int(row.get("leavers") or 0),
        }
        for row in rows
    ]


def _parse_weekly_excel(excel_path: Path) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook["Sheet2"]
    main_rows: list[dict] = []
    current: dict | None = None
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if values[0] == "合计":
            break
        if values[1] is not None:
            current = {
                "business_unit": values[1],
                "headcount": values[2],
                "cnt_formal": values[3],
                "cnt_intern": values[4],
                "cnt_labor": values[5],
                "joiners": values[6],
                "leavers": values[7],
                "top3_projects": [],
            }
            main_rows.append(current)
        if current is not None and values[8] is not None:
            current["top3_projects"].append(
                {"name": values[8], "count": values[9]}
            )

    sheet1 = workbook["Sheet1"]
    cc_rows = [
        {
            "cost_center": values[0],
            "project": values[1],
            "headcount": values[2],
            "joiners": values[3],
            "leavers": values[4],
        }
        for values in sheet1.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values[:5])
    ]
    return main_rows, cc_rows


def _verify_excel(snapshot: PreviewSnapshot, excel_path: Path) -> None:
    if snapshot.report_kind == "daily":
        projection = _daily_excel_projection(snapshot, excel_path)
        if projection["expected_rows"] != projection["actual_rows"]:
            raise PublicationFailed("generated daily Excel does not match preview rows")
        if projection["expected_tenure"] != projection["actual_tenure"]:
            raise PublicationFailed("generated daily Excel tenure does not match preview")
        return

    actual_main, actual_cc = _parse_weekly_excel(excel_path)
    if _weekly_main_projection(snapshot.main_rows) != _weekly_main_projection(
        actual_main
    ):
        raise PublicationFailed("generated weekly Excel Sheet2 does not match preview")
    if _weekly_cc_projection(snapshot.cc_rows) != _weekly_cc_projection(actual_cc):
        raise PublicationFailed("generated weekly Excel Sheet1 does not match preview")


def _write_manifest(
    db: Session,
    run: ReportRun,
    snapshot: PreviewSnapshot,
    paths: Mapping[str, Path],
    staging_dir: Path,
) -> tuple[Path, dict]:
    sources = db.scalars(
        select(RunSource)
        .where(RunSource.run_id == run.id, RunSource.is_deleted == 0)
        .order_by(RunSource.source_type)
    ).all()
    manifest = {
        "run_id": run.id,
        "report_kind": snapshot.report_kind,
        "period_start": snapshot.period_start.isoformat(),
        "period_end": snapshot.period_end.isoformat(),
        "rule_version": snapshot.rule_version,
        "snapshot_hash": snapshot.snapshot_hash,
        "source_hashes": {
            source.source_type: source.sha256 for source in sources
        },
        "artifacts": {
            kind: {
                "filename": path.name,
                "sha256": _sha256(path),
                "size_bytes": path.stat().st_size,
            }
            for kind, path in sorted(paths.items())
        },
    }
    manifest_path = staging_dir / _ARTIFACT_FILENAMES["manifest"]
    manifest_path.write_text(encode_json_text(manifest), encoding="utf-8")
    return manifest_path, manifest


def _mark_pre_metadata_failure(
    db: Session,
    attempt_id: str,
    run_id: str,
    report_kind: str,
    exc: BaseException,
) -> None:
    db.rollback()
    attempt = db.get(PublicationAttempt, attempt_id)
    target = db.scalar(
        select(RunReportTarget).where(
            RunReportTarget.run_id == run_id,
            RunReportTarget.report_kind == report_kind,
            RunReportTarget.is_deleted == 0,
        )
    )
    if attempt is not None:
        attempt.status = "failed"
        attempt.error_message_redacted = _redacted_error(exc)
        attempt.completed_at = _now()
    if target is not None and target.status != TargetStatus.published.value:
        target.status = TargetStatus.failed.value
        target.error_code = "publication_failed"
        target.error_message_redacted = _redacted_error(exc)
    db.commit()


def _publish_one(
    db: Session,
    run: ReportRun,
    report_kind: str,
    operator_ref: str,
    *,
    bundle: FactBundle | None,
    period: tuple[date, date] | None,
    template_path: str | None,
    output_root: Path,
) -> PublishedReport:
    started = time.perf_counter()

    def checkpoint(stage: str) -> None:
        log.info(
            "publication stage run=%s kind=%s stage=%s total_seconds=%.3f",
            run.id,
            report_kind,
            stage,
            time.perf_counter() - started,
        )

    target = run_repo.ensure_report_targets(db, run.id, (report_kind,))[0]
    existing_report = db.scalar(
        select(PublishedReport).where(
            PublishedReport.run_id == run.id,
            PublishedReport.report_kind == report_kind,
            PublishedReport.is_deleted == 0,
        )
    )
    expected_hash = target.preview_hash
    if not expected_hash:
        raise PreviewRequired(f"{report_kind} must be previewed before publication")
    # 同 Run 已发布同 snapshot → 幂等返回已发布的报告
    if existing_report is not None:
        if existing_report.snapshot_hash == expected_hash:
            log.info("%s was already published with the same snapshot, returning existing report", report_kind)
            return existing_report
        raise PublicationFailed(
            f"{report_kind} was already published for this run; create a revision run"
        )
    week_start = period[0] if period else None
    week_end = period[1] if period else None
    snapshot = build_preview(
        db,
        run.id,
        report_kind,
        bundle=bundle,
        week_start=week_start,
        week_end=week_end,
        persist_preview_hash=False,
        reuse_published_snapshot=False,
    )
    checkpoint("preview")
    if snapshot.snapshot_hash != expected_hash:
        target = run_repo.ensure_report_targets(db, run.id, (report_kind,))[0]
        target.preview_hash = None
        target.status = TargetStatus.needs_review.value
        target.error_code = "preview_changed"
        target.error_message_redacted = "facts or validation changed after preview"
        db.commit()
        raise PreviewChanged(f"{report_kind} preview changed; preview again")
    if not snapshot.publishable:
        raise PublicationBlocked(f"{report_kind} validations are not publishable")

    attempt = PublicationAttempt(
        run_id=run.id,
        report_kind=report_kind,
        status="staging",
        staging_path="pending",
        final_path="pending",
    )
    db.add(attempt)
    db.flush()
    attempt_id = attempt.id
    period_token = (
        snapshot.period_end.isoformat()
        if report_kind == "daily"
        else f"{snapshot.period_start.isoformat()}_{snapshot.period_end.isoformat()}"
    )
    staging_dir = (
        output_root / ".publication-staging" / attempt_id
    ).resolve()
    final_dir = (
        output_root / "published" / report_kind / period_token / attempt_id
    ).resolve()
    attempt.staging_path = str(staging_dir)
    attempt.final_path = str(final_dir)
    target.status = TargetStatus.publishing.value
    target.error_code = None
    target.error_message_redacted = None
    db.commit()
    checkpoint("staging_committed")

    try:
        paths = _write_artifacts(
            db,
            snapshot,
            staging_dir,
            template_path=template_path,
            output_root=output_root,
        )
        checkpoint("artifacts_written")
        _verify_excel(snapshot, paths["excel"])
        checkpoint("excel_verified")
        manifest_path, manifest = _write_manifest(
            db, run, snapshot, paths, staging_dir
        )
        checkpoint("manifest_written")
        paths = {**paths, "manifest": manifest_path}
    except Exception as exc:
        shutil.rmtree(staging_dir, ignore_errors=True)
        _mark_pre_metadata_failure(db, attempt_id, run.id, report_kind, exc)
        raise PublicationFailed(
            f"publication failed: {_redacted_error(exc)}"
        ) from exc

    try:
        published_at = _now()
        publication_repo.supersede_current(
            db,
            report_kind,
            snapshot.period_start,
            snapshot.period_end,
            published_at,
        )
        report = PublishedReport(
            run_id=run.id,
            report_kind=report_kind,
            period_start=snapshot.period_start,
            period_end=snapshot.period_end,
            version=publication_repo.next_version(
                db, report_kind, snapshot.period_start, snapshot.period_end
            ),
            is_current=True,
            snapshot_json=snapshot.snapshot_json,
            snapshot_hash=snapshot.snapshot_hash,
            baseline_report_id=run.baseline_report_id,
            published_by=operator_ref,
            published_at=published_at,
        )
        db.add(report)
        db.flush()
        for kind, path in paths.items():
            db.add(
                ReportArtifact(
                    report_id=report.id,
                    artifact_kind=kind,
                    protected_path=str(final_dir / path.name),
                    sha256=_sha256(path),
                    size_bytes=path.stat().st_size,
                )
            )
        if report_kind == "daily":
            report_repo.save_daily(
                db,
                snapshot.period_end,
                dict(snapshot.calculation_context["rows"]),
                commit=False,
            )
            report_repo.save_tenure_snapshot(
                db,
                snapshot.period_end,
                list(snapshot.calculation_context["tenure"]["rows"]),
                commit=False,
            )
        else:
            report_repo.save_weekly(
                db,
                snapshot.period_start,
                snapshot.period_end,
                list(snapshot.main_rows),
                commit=False,
            )
        target = run_repo.ensure_report_targets(db, run.id, (report_kind,))[0]
        target.status = TargetStatus.published.value
        target.published_report_id = report.id
        attempt = db.get(PublicationAttempt, attempt_id)
        attempt.status = "metadata_committed"
        attempt.report_id = report.id
        attempt.manifest_json = encode_json_text(manifest)
        db.commit()
        checkpoint("metadata_committed")
    except Exception as exc:
        db.rollback()
        try:
            db.expire_all()
            persisted_attempt = db.get(PublicationAttempt, attempt_id)
        except Exception as lookup_exc:
            raise PublicationRecoveryRequired(
                "publication commit outcome is unknown; preserve staging for recovery"
            ) from lookup_exc
        if (
            persisted_attempt is not None
            and persisted_attempt.status == "metadata_committed"
            and persisted_attempt.report_id is not None
        ):
            raise PublicationRecoveryRequired(
                "publication metadata committed; artifact move requires recovery"
            ) from exc
        shutil.rmtree(staging_dir, ignore_errors=True)
        _mark_pre_metadata_failure(db, attempt_id, run.id, report_kind, exc)
        raise PublicationFailed(
            f"publication metadata failed: {_redacted_error(exc)}"
        ) from exc

    try:
        final_dir.parent.mkdir(parents=True, exist_ok=True)
        os.replace(staging_dir, final_dir)
        checkpoint("artifacts_moved")
    except Exception as exc:
        db.rollback()
        attempt = db.get(PublicationAttempt, attempt_id)
        if attempt is not None:
            attempt.error_message_redacted = _redacted_error(exc)
            db.commit()
        raise PublicationRecoveryRequired(
            "publication metadata committed; artifact move requires recovery: "
            f"{_redacted_error(exc)}"
        ) from exc

    attempt = db.get(PublicationAttempt, attempt_id)
    attempt.status = "completed"
    attempt.completed_at = _now()
    db.commit()
    checkpoint("completed")
    return db.get(PublishedReport, report.id)


def publish(
    db: Session,
    run_id: str,
    report_kinds: Sequence[str],
    operator_ref: str,
    *,
    bundles: Mapping[str, FactBundle] | None = None,
    periods: Mapping[str, tuple[date, date]] | None = None,
    template_paths: Mapping[str, str] | None = None,
    output_dir: str | Path | None = None,
) -> list[PublishedReport]:
    kinds = tuple(dict.fromkeys(report_kinds))
    if not kinds or set(kinds) - {"daily", "weekly"}:
        raise ValueError("report_kinds must contain daily and/or weekly")
    operator = str(operator_ref or "").strip()
    if not operator:
        raise ValueError("operator_ref is required")
    run = db.get(ReportRun, run_id)
    if run is None or run.is_deleted:
        raise LookupError(f"Run {run_id} was not found")
    output_root = Path(output_dir or settings.output_dir).resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    bundles = bundles or {}
    periods = periods or {}
    template_paths = template_paths or {}
    return [
        _publish_one(
            db,
            run,
            kind,
            operator,
            bundle=bundles.get(kind),
            period=periods.get(kind),
            template_path=template_paths.get(kind),
            output_root=output_root,
        )
        for kind in kinds
    ]


def recover_publication_attempts(
    db: Session, output_dir: str | Path | None = None
) -> list[str]:
    """Finish committed moves or remove abandoned pre-metadata staging trees."""
    _ = Path(output_dir or settings.output_dir).resolve()
    attempts = db.scalars(
        select(PublicationAttempt).where(
            PublicationAttempt.status.in_(("staging", "metadata_committed")),
            PublicationAttempt.is_deleted == 0,
        )
    ).all()
    recovered: list[str] = []
    for attempt in attempts:
        staging = Path(attempt.staging_path)
        final = Path(attempt.final_path)
        if attempt.status == "staging":
            shutil.rmtree(staging, ignore_errors=True)
            attempt.status = "failed"
            attempt.error_message_redacted = "abandoned before metadata commit"
            attempt.completed_at = _now()
            recovered.append(attempt.id)
            continue
        try:
            if not final.exists() and staging.exists():
                final.parent.mkdir(parents=True, exist_ok=True)
                os.replace(staging, final)
            if not final.is_dir():
                raise FileNotFoundError("staging and final artifact directories are absent")
            attempt.status = "completed"
            attempt.completed_at = _now()
            attempt.error_message_redacted = None
            recovered.append(attempt.id)
        except Exception as exc:
            attempt.error_message_redacted = _redacted_error(exc)
    db.commit()
    return recovered
